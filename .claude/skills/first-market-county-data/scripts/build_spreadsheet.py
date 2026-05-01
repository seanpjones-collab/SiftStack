"""
Build a First-to-Market Data Sources XLSX using the canonical template.

Canonical template derived from Stark/Summit/Cuyahoga example reports.

Usage:
    python build_spreadsheet.py <input.json> <output.xlsx>

Input JSON schema:
{
  "county": "Cuyahoga",
  "state": "Ohio",
  "subtitle": "Primary city: Cleveland | Judicial Foreclosure State | Researched April 2026",
  "rows": [
    {
      "priority": "A",                    # A, B, or C
      "data_type": "Probate / Heirship",
      "office": "Cuyahoga County Probate Court",
      "address": "1 Lakeside Ave, Cleveland OH 44113",
      "phone": "(216) 443-8764",
      "portal": "https://probate.cuyahogacounty.gov/pa/",
      "foia": "probate@cuyahogacounty.us",
      "difficulty": "Medium",             # Low, Low-Medium, Medium, High
      "notes": "Long notes field...",
      "freshness": "Daily / real-time docket"
    },
    ...
  ],
  "county_notes": [
    "Ohio is a JUDICIAL FORECLOSURE state...",
    "Cleveland code violations available via Socrata API...",
  ]
}
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- Canonical template constants ---
TITLE_BG = "1F3864"
SUBTITLE_BG = "2E75B6"
HEADER_BG = "1F3864"
WHITE = "FFFFFF"
FONT_NAME = "Arial"

PRIORITY_FILLS = {
    "A": ("C6EFCE", "A9D18E"),  # green stripes (light/dark)
    "B": ("FFEB9C", "FFD966"),  # yellow stripes
    "C": ("FCE4D6", "F4B183"),  # orange stripes
}

COLUMNS = [
    ("Priority", 8),
    ("Data Type", 22),
    ("Office / Division", 30),
    ("Address", 40),
    ("Phone", 26),
    ("Portal URL", 48),
    ("FOIA / Contact Email", 42),
    ("Difficulty", 12),
    ("Notes", 55),
    ("Data Freshness", 28),
]

ROW_KEYS = ["priority", "data_type", "office", "address", "phone",
            "portal", "foia", "difficulty", "notes", "freshness"]


def build(data: dict, output_path: str) -> None:
    county = data["county"]
    state = data.get("state", "")
    subtitle = data.get("subtitle", "")
    rows = data["rows"]
    notes = data.get("county_notes", [])

    wb = Workbook()
    ws = wb.active
    # Excel sheet names capped at 31 chars — keep it short and consistent.
    ws.title = f"{county} County Data Sources"[:31]

    n_cols = len(COLUMNS)
    last_col = get_column_letter(n_cols)

    # Row 1 — Title
    ws.cell(row=1, column=1,
            value=f"{county} County, {state} — First-to-Market Distress Data Sources")
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.font = Font(name=FONT_NAME, bold=True, size=16, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Row 2 — Subtitle
    ws.cell(row=2, column=1, value=subtitle)
    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws["A2"]
    sub_cell.font = Font(name=FONT_NAME, size=10, color=WHITE)
    sub_cell.fill = PatternFill("solid", fgColor=SUBTITLE_BG)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 19.5

    # Row 3 — Headers
    for i, (header, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=3, column=i, value=header)
        c.font = Font(name=FONT_NAME, bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[3].height = 31.5

    # Data rows with alternating stripe fills per priority group
    priority_counter = {"A": 0, "B": 0, "C": 0}
    for row_idx, row_data in enumerate(rows, start=4):
        pri = row_data.get("priority", "C")
        stripe_idx = priority_counter.get(pri, 0) % 2
        fill_color = PRIORITY_FILLS.get(pri, PRIORITY_FILLS["C"])[stripe_idx]
        priority_counter[pri] = priority_counter.get(pri, 0) + 1

        values = [row_data.get(k, "") for k in ROW_KEYS]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 1:  # Priority bold
                c.font = Font(name=FONT_NAME, bold=True, size=10)
            else:
                c.font = Font(name=FONT_NAME, size=10)
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{last_col}{ws.max_row}"

    # --- Sheet 2: Legend & Notes ---
    ws2 = wb.create_sheet("Legend & Notes")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 80

    def _section_header(row, title):
        c = ws2.cell(row=row, column=1, value=title)
        c.font = Font(name=FONT_NAME, bold=True, size=11)

    legend_rows = [
        ("Priority", "Color", "Description"),
        ("A", "Green", "Core first-to-market lists — highest ROI for wholesaling (Probate, Foreclosure, Tax Sale, Tax Delinquency)"),
        ("B", "Yellow", "Standard distress lists — code violations, condemned, mechanic's liens, IRS/state tax liens"),
        ("C", "Orange", "Extended lists — HOA liens, utility shut-offs, permits, environmental, fire/storm, Medicaid, evictions, etc."),
        (None, None, None),
        ("Difficulty", "Meaning", None),
        ("Low", "Online portal with bulk export; minimal barriers", None),
        ("Low-Medium", "Online searchable but bulk export limited; may need account or FOIA", None),
        ("Medium", "Account registration, limited export, or in-person pickup required", None),
        ("High", "FOIA request, physical visit, affidavit, or paywall required", None),
        (None, None, None),
    ]
    for r_idx, r in enumerate(legend_rows, 1):
        for c_idx, v in enumerate(r, 1):
            if v is not None:
                cell = ws2.cell(row=r_idx, column=c_idx, value=v)
                if r_idx == 1 or (r_idx == 6 and c_idx <= 2):
                    cell.font = Font(name=FONT_NAME, bold=True)
                else:
                    cell.font = Font(name=FONT_NAME)

    # Priority color swatches on legend rows
    legend_swatches = {2: PRIORITY_FILLS["A"][0], 3: PRIORITY_FILLS["B"][0], 4: PRIORITY_FILLS["C"][0]}
    for r_idx, color in legend_swatches.items():
        for c_idx in range(1, 4):
            ws2.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor=color)

    # County-specific notes
    start_row = len(legend_rows) + 1
    _section_header(start_row, f"{county} County — Key Notes")
    for i, note in enumerate(notes, 1):
        cell = ws2.cell(row=start_row + i, column=1, value=f"{i}. {note}")
        ws2.merge_cells(start_row=start_row + i, start_column=1,
                        end_row=start_row + i, end_column=3)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(f"Wrote {output_path}  ({len(rows)} data rows)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)
    build(data, sys.argv[2])
