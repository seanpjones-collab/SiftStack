#!/usr/bin/env python3
"""
Generate comprehensive Excel breakdown for real estate comping analysis.

Usage:
    python generate_excel_report.py <output_path> <json_data_path>
    
    OR import and call generate_excel_report() directly with data dict.

The JSON data file should contain:
{
    "subject_property": {...},
    "comps": [...],
    "market_overview": {...},
    "arv_calculation": {...},
    "adjustments_applied": [...],
    "sources": [...],
    "recommendations": [...]
}
"""

import sys
import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


def create_header_style():
    """Create consistent header styling."""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_data_style():
    """Create consistent data cell styling."""
    return {
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_currency_style():
    """Create currency formatting style."""
    style = create_data_style()
    style['alignment'] = Alignment(horizontal='right', vertical='center')
    return style


def create_section_header_style():
    """Create section header styling."""
    return {
        'font': Font(bold=True, size=14, color='2F5496'),
        'alignment': Alignment(horizontal='left', vertical='center')
    }


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def auto_fit_columns(ws, min_width=10, max_width=50):
    """Auto-fit column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_excel_report(data: dict, output_path: str) -> str:
    """
    Generate comprehensive Excel report from comping analysis data.
    
    Args:
        data: Dictionary containing all analysis data
        output_path: Path for output Excel file
        
    Returns:
        Path to generated Excel file
    """
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    summary_ws = wb.create_sheet("Executive Summary")
    subject_ws = wb.create_sheet("Subject Property")
    comps_ws = wb.create_sheet("Comparable Sales")
    adjustments_ws = wb.create_sheet("Adjustments Detail")
    market_ws = wb.create_sheet("Market Analysis")
    arv_ws = wb.create_sheet("ARV Calculation")
    sources_ws = wb.create_sheet("Sources & Notes")
    
    header_style = create_header_style()
    data_style = create_data_style()
    currency_style = create_currency_style()
    section_style = create_section_header_style()
    
    # ========== EXECUTIVE SUMMARY SHEET ==========
    ws = summary_ws
    ws.merge_cells('A1:F1')
    ws['A1'] = "PROPERTY VALUATION REPORT - EXECUTIVE SUMMARY"
    ws['A1'].font = Font(bold=True, size=18, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A3'].font = Font(italic=True, size=10)
    
    # Subject property quick view
    ws['A5'] = "SUBJECT PROPERTY"
    apply_style(ws['A5'], section_style)
    
    subject = data.get('subject_property', {})
    row = 6
    summary_fields = [
        ('Address', subject.get('address', 'N/A')),
        ('City, State, ZIP', f"{subject.get('city', '')}, {subject.get('state', '')} {subject.get('zip', '')}"),
        ('Property Type', subject.get('property_type', 'N/A')),
        ('Living Area (GLA)', f"{subject.get('gla', 'N/A'):,} sqft" if isinstance(subject.get('gla'), (int, float)) else 'N/A'),
        ('Beds / Baths', f"{subject.get('beds', 'N/A')} / {subject.get('baths', 'N/A')}"),
        ('Year Built', subject.get('year_built', 'N/A')),
    ]
    
    for label, value in summary_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # ARV Summary
    row += 1
    ws[f'A{row}'] = "ARV ANALYSIS RESULTS"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    arv = data.get('arv_calculation', {})
    arv_fields = [
        ('Final ARV', f"${arv.get('final_arv', 0):,.0f}" if arv.get('final_arv') else 'N/A'),
        ('ARV Range (Low)', f"${arv.get('arv_low', 0):,.0f}" if arv.get('arv_low') else 'N/A'),
        ('ARV Range (High)', f"${arv.get('arv_high', 0):,.0f}" if arv.get('arv_high') else 'N/A'),
        ('Confidence Level', arv.get('confidence_level', 'N/A')),
        ('Renovated PPSF', f"${arv.get('renovated_ppsf', arv.get('post_reno_ppsf', 0)):,.2f}" if arv.get('renovated_ppsf', arv.get('post_reno_ppsf')) else 'N/A'),
    ]
    
    for label, value in arv_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if 'ARV' in label:
            ws[f'B{row}'].font = Font(bold=True, size=12, color='2F5496')
        row += 1
    
    # Market snapshot
    row += 1
    ws[f'A{row}'] = "MARKET SNAPSHOT"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    market = data.get('market_overview', {})
    market_fields = [
        ('Market Phase', market.get('market_phase', 'N/A')),
        ('Median Price', f"${market.get('median_price', 0):,.0f}" if market.get('median_price') else 'N/A'),
        ('Median PPSF', f"${market.get('median_ppsf', 0):,.2f}" if market.get('median_ppsf') else 'N/A'),
        ('Avg Days on Market', market.get('avg_dom', 'N/A')),
        ('Sale-to-List Ratio', f"{market.get('sale_to_list_ratio', 0):.1%}" if market.get('sale_to_list_ratio') else 'N/A'),
    ]
    
    for label, value in market_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    auto_fit_columns(ws)
    
    # ========== SUBJECT PROPERTY SHEET ==========
    ws = subject_ws
    ws['A1'] = "SUBJECT PROPERTY DETAILS"
    ws['A1'].font = Font(bold=True, size=16, color='2F5496')
    
    row = 3
    all_subject_fields = [
        ('Basic Information', None),
        ('Address', subject.get('address', 'N/A')),
        ('City', subject.get('city', 'N/A')),
        ('State', subject.get('state', 'N/A')),
        ('ZIP Code', subject.get('zip', 'N/A')),
        ('County', subject.get('county', 'N/A')),
        ('Subdivision', subject.get('subdivision', 'N/A')),
        ('', None),
        ('Property Characteristics', None),
        ('Property Type', subject.get('property_type', 'N/A')),
        ('Living Area (GLA)', f"{subject.get('gla', 'N/A'):,} sqft" if isinstance(subject.get('gla'), (int, float)) else 'N/A'),
        ('Lot Size', f"{subject.get('lot_size', 'N/A'):,} sqft" if isinstance(subject.get('lot_size'), (int, float)) else 'N/A'),
        ('Bedrooms', subject.get('beds', 'N/A')),
        ('Bathrooms', subject.get('baths', 'N/A')),
        ('Year Built', subject.get('year_built', 'N/A')),
        ('Stories', subject.get('stories', 'N/A')),
        ('Garage', subject.get('garage', 'N/A')),
        ('Pool', subject.get('pool', 'N/A')),
        ('Basement', subject.get('basement', 'N/A')),
        ('', None),
        ('Current Condition', None),
        ('Condition Rating', subject.get('condition', 'N/A')),
        ('', None),
        ('Ownership & Legal', None),
        ('Owner Name', subject.get('owner_name', 'N/A')),
        ('Owner Type', subject.get('owner_type', 'N/A')),
        ('Zoning', subject.get('zoning', 'N/A')),
        ('HOA', subject.get('hoa', 'N/A')),
        ('Flood Zone', subject.get('flood_zone', 'N/A')),
    ]
    
    for label, value in all_subject_fields:
        if value is None and label:
            ws[f'A{row}'] = label
            apply_style(ws[f'A{row}'], section_style)
        elif label:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
        row += 1
    
    auto_fit_columns(ws)
    
    # ========== COMPARABLE SALES SHEET ==========
    ws = comps_ws
    ws['A1'] = "COMPARABLE SALES ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16, color='2F5496')
    
    comps = data.get('comps', [])
    
    headers = ['#', 'Address', 'Sale Date', 'Sale Price', 'GLA', 'PPSF', 'Beds', 'Baths', 
               'Year Built', 'Condition', 'Distance', 'Adjustments', 'Adjusted Value']
    
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, header_style)
    
    for i, comp in enumerate(comps, 1):
        row += 1
        values = [
            i,
            comp.get('address', 'N/A'),
            comp.get('sale_date', 'N/A'),
            f"${comp.get('sale_price', 0):,.0f}" if comp.get('sale_price') else 'N/A',
            f"{comp.get('gla', 0):,}" if comp.get('gla') else 'N/A',
            f"${comp.get('ppsf', 0):,.2f}" if comp.get('ppsf') else 'N/A',
            comp.get('beds', 'N/A'),
            comp.get('baths', 'N/A'),
            comp.get('year_built', 'N/A'),
            comp.get('condition', 'N/A'),
            f"{comp.get('distance', 'N/A')} mi" if comp.get('distance') else 'N/A',
            f"${comp.get('total_adjustments', 0):+,.0f}" if comp.get('total_adjustments') else '$0',
            f"${comp.get('adjusted_value', 0):,.0f}" if comp.get('adjusted_value') else 'N/A',
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_style(cell, data_style)
    
    # Add bucket analysis section
    row += 3
    ws[f'A{row}'] = "BUCKET ANALYSIS (Two-Bucket Method)"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    bucket_headers = ['Bucket', 'Description', 'Comp Count', 'Median PPSF', 'Avg PPSF']
    for col, header in enumerate(bucket_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, header_style)
    
    buckets = data.get('bucket_analysis', {})
    row += 1
    unreno = buckets.get('unrenovated', {})
    ws.cell(row=row, column=1, value='A - Unrenovated')
    ws.cell(row=row, column=2, value='Dated/original condition')
    ws.cell(row=row, column=3, value=unreno.get('count', 0))
    ws.cell(row=row, column=4, value=f"${unreno.get('median_ppsf', 0):,.2f}" if unreno.get('median_ppsf') else 'N/A')
    ws.cell(row=row, column=5, value=f"${unreno.get('avg_ppsf', 0):,.2f}" if unreno.get('avg_ppsf') else 'N/A')
    
    row += 1
    reno = buckets.get('renovated', {})
    ws.cell(row=row, column=1, value='B - Renovated')
    ws.cell(row=row, column=2, value='Fully updated/flipped')
    ws.cell(row=row, column=3, value=reno.get('count', 0))
    ws.cell(row=row, column=4, value=f"${reno.get('median_ppsf', 0):,.2f}" if reno.get('median_ppsf') else 'N/A')
    ws.cell(row=row, column=5, value=f"${reno.get('avg_ppsf', 0):,.2f}" if reno.get('avg_ppsf') else 'N/A')
    
    row += 2
    reno_premium = buckets.get('market_premium_pct', buckets.get('renovation_premium_pct', 0))
    ws[f'A{row}'] = "Market Premium (Bucket Spread):"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f"{reno_premium:.1f}%" if reno_premium else 'N/A'
    ws[f'B{row}'].font = Font(bold=True, color='2F5496')
    
    auto_fit_columns(ws)
    
    # ========== ADJUSTMENTS DETAIL SHEET ==========
    ws = adjustments_ws
    ws['A1'] = "ADJUSTMENTS BREAKDOWN"
    ws['A1'].font = Font(bold=True, size=16, color='2F5496')
    
    adj_headers = ['Comp #', 'Address', 'Adjustment Type', 'Reason', 'Amount']
    row = 3
    for col, header in enumerate(adj_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        apply_style(cell, header_style)
    
    adjustments = data.get('adjustments_applied', [])
    for adj in adjustments:
        row += 1
        values = [
            adj.get('comp_number', 'N/A'),
            adj.get('comp_address', 'N/A'),
            adj.get('adjustment_type', 'N/A'),
            adj.get('reason', 'N/A'),
            f"${adj.get('amount', 0):+,.0f}" if adj.get('amount') else '$0',
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            apply_style(cell, data_style)
    
    # Adjustment summary
    row += 3
    ws[f'A{row}'] = "ADJUSTMENT SUMMARY BY TYPE"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    adj_summary = data.get('adjustment_summary', {})
    for adj_type, total in adj_summary.items():
        ws[f'A{row}'] = adj_type
        ws[f'B{row}'] = f"${total:+,.0f}" if isinstance(total, (int, float)) else total
        row += 1
    
    auto_fit_columns(ws)
    
    # ========== MARKET ANALYSIS SHEET ==========
    ws = market_ws
    ws['A1'] = "MARKET ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16, color='2F5496')
    
    row = 3
    ws[f'A{row}'] = "MARKET METRICS"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    market_metrics = [
        ('Market Phase', market.get('market_phase', 'N/A')),
        ('Median Sale Price', f"${market.get('median_price', 0):,.0f}" if market.get('median_price') else 'N/A'),
        ('Median Price Per Sqft', f"${market.get('median_ppsf', 0):,.2f}" if market.get('median_ppsf') else 'N/A'),
        ('Average Days on Market', market.get('avg_dom', 'N/A')),
        ('Sale-to-List Ratio', f"{market.get('sale_to_list_ratio', 0):.1%}" if market.get('sale_to_list_ratio') else 'N/A'),
        ('% Selling Over List', f"{market.get('pct_over_list', 0):.1%}" if market.get('pct_over_list') else 'N/A'),
        ('Active Listings', market.get('active_count', 'N/A')),
        ('Pending Listings', market.get('pending_count', 'N/A')),
        ('Months of Inventory', market.get('months_inventory', 'N/A')),
    ]
    
    for label, value in market_metrics:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    ws[f'A{row}'] = "MARKET TRENDS"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    trends = market.get('trends', {})
    trend_items = [
        ('Price Trend (3 mo)', trends.get('price_trend_3mo', 'N/A')),
        ('Price Trend (6 mo)', trends.get('price_trend_6mo', 'N/A')),
        ('DOM Trend', trends.get('dom_trend', 'N/A')),
        ('Inventory Trend', trends.get('inventory_trend', 'N/A')),
    ]
    
    for label, value in trend_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    ws[f'A{row}'] = "MARKET NOTES"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    notes = market.get('notes', [])
    for note in notes:
        ws[f'A{row}'] = f"• {note}"
        row += 1
    
    auto_fit_columns(ws)
    
    # ========== ARV CALCULATION SHEET ==========
    ws = arv_ws
    ws['A1'] = "ARV CALCULATION BREAKDOWN"
    ws['A1'].font = Font(bold=True, size=16, color='2F5496')
    
    row = 3
    ws[f'A{row}'] = "STEP-BY-STEP ARV CALCULATION"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    arv_steps = [
        ('Step 1: Base PPSF (Unrenovated)', f"${arv.get('base_ppsf', 0):,.2f}" if arv.get('base_ppsf') else 'N/A'),
        ('Step 2: Market Premium (Bucket Spread)', f"{arv.get('market_premium_pct', arv.get('renovation_premium_pct', 0)):.1f}%" if arv.get('market_premium_pct', arv.get('renovation_premium_pct')) else 'N/A'),
        ('Step 3: Renovated PPSF', f"${arv.get('renovated_ppsf', arv.get('post_reno_ppsf', 0)):,.2f}" if arv.get('renovated_ppsf', arv.get('post_reno_ppsf')) else 'N/A'),
        ('Step 4: Market Sentiment Adj', f"{arv.get('sentiment_adjustment_pct', 0):+.1f}%" if arv.get('sentiment_adjustment_pct') else '0%'),
        ('Step 5: Adjusted PPSF', f"${arv.get('adjusted_ppsf', 0):,.2f}" if arv.get('adjusted_ppsf') else 'N/A'),
        ('Step 6: Subject GLA', f"{arv.get('subject_gla', 0):,} sqft" if arv.get('subject_gla') else 'N/A'),
        ('Step 7: Base ARV (PPSF × GLA)', f"${arv.get('base_arv', 0):,.0f}" if arv.get('base_arv') else 'N/A'),
        ('Step 8: Feature Adjustments', f"${arv.get('feature_adjustments', 0):+,.0f}" if arv.get('feature_adjustments') else '$0'),
        ('Step 9: FINAL ARV', f"${arv.get('final_arv', 0):,.0f}" if arv.get('final_arv') else 'N/A'),
    ]
    
    for label, value in arv_steps:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        if 'FINAL' in label:
            ws[f'B{row}'].font = Font(bold=True, size=14, color='2F5496')
        row += 1
    
    row += 1
    ws[f'A{row}'] = "CONFIDENCE ANALYSIS"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    confidence_items = [
        ('Confidence Level', arv.get('confidence_level', 'N/A')),
        ('Confidence Band', f"±{arv.get('confidence_band_pct', 0):.1f}%" if arv.get('confidence_band_pct') else 'N/A'),
        ('ARV Low', f"${arv.get('arv_low', 0):,.0f}" if arv.get('arv_low') else 'N/A'),
        ('ARV High', f"${arv.get('arv_high', 0):,.0f}" if arv.get('arv_high') else 'N/A'),
    ]
    
    for label, value in confidence_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    auto_fit_columns(ws)
    
    # ========== SOURCES & NOTES SHEET ==========
    ws = sources_ws
    ws['A1'] = "SOURCES & NOTES"
    ws['A1'].font = Font(bold=True, size=16, color='2F5496')
    
    row = 3
    ws[f'A{row}'] = "DATA SOURCES"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    sources = data.get('sources', [])
    for source in sources:
        ws[f'A{row}'] = f"• {source}"
        row += 1
    
    row += 1
    ws[f'A{row}'] = "SEARCH PARAMETERS"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    params = data.get('search_parameters', {})
    param_items = [
        ('Time Window', params.get('time_window', 'N/A')),
        ('Search Radius', params.get('radius', 'N/A')),
        ('GLA Range', params.get('gla_range', 'N/A')),
        ('Year Built Range', params.get('year_range', 'N/A')),
        ('Subdivision Constraint', params.get('subdivision', 'N/A')),
    ]
    
    for label, value in param_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    ws[f'A{row}'] = "RECOMMENDATIONS"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    recommendations = data.get('recommendations', [])
    for rec in recommendations:
        ws[f'A{row}'] = f"• {rec}"
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 1
    ws[f'A{row}'] = "CAVEATS & DISCLAIMERS"
    apply_style(ws[f'A{row}'], section_style)
    row += 1
    
    caveats = data.get('caveats', [
        "This analysis mimics an appraisal process but is NOT a formal appraisal.",
        "Values are estimates based on available market data and may vary.",
        "Professional verification recommended before making investment decisions.",
        "Market conditions can change rapidly; re-verify before closing."
    ])
    for caveat in caveats:
        ws[f'A{row}'] = f"• {caveat}"
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    auto_fit_columns(ws)
    
    # Save workbook
    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_excel_report.py <output_path> <json_data_path>")
        print("\nExample: python generate_excel_report.py report.xlsx data.json")
        sys.exit(1)
    
    output_path = sys.argv[1]
    json_path = sys.argv[2]
    
    with open(json_path, 'r') as f:
        data = json.load(f)
    
    result = generate_excel_report(data, output_path)
    print(f"Excel report generated: {result}")


if __name__ == "__main__":
    main()
