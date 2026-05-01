#!/usr/bin/env python3
"""
Rehab Cost Estimator — Excel Workbook Generator
Produces a multi-sheet Excel workbook with single-estimate (no Low/Mid/High range).

Sheets:
1. Summary — Executive overview with both strategies side-by-side
2. Full Rehab Estimate — Itemized line items with single estimate
3. Wholetail Estimate — Itemized line items with single estimate
4. Condition Assessment — Property condition analysis
5. Deal Analyzer — Investment math (ARV, purchase, profit, ROI, 75% MAO)
6. Budget Tracker — Empty template for actual cost tracking
7. Material Specs — Recommended materials with finish tier specs
8. Project Checklist — Construction timeline template

Usage:
    python generate_rehab_excel.py <output_path> <json_data_path>
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter


# ── Styles ──────────────────────────────────────────────────────────────────

BLUE = "2F5496"
DARK_BLUE = "1F3864"
GREEN = "548235"
RED = "C00000"
ORANGE = "ED7D31"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
dark_header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
subtitle_font = Font(name="Calibri", bold=True, size=12, color=BLUE)
section_font = Font(name="Calibri", bold=True, size=11, color=BLUE)
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
currency_format = '#,##0'
currency_decimal_format = '#,##0.00'
pct_format = '0.0%'
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
highlight_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col, fill=None):
    if fill is None:
        fill = header_fill
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_cell(cell, is_currency=False, is_pct=False, alt_row=False):
    cell.font = normal_font
    cell.border = thin_border
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if is_currency:
        cell.number_format = currency_format
        cell.alignment = Alignment(horizontal="right", vertical="center")
    if is_pct:
        cell.number_format = pct_format
        cell.alignment = Alignment(horizontal="right", vertical="center")
    if alt_row:
        cell.fill = alt_fill


def auto_fit_columns(ws, min_width=10, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def fmt_currency(val):
    if val is None or val == 0:
        return "$0"
    return f"${val:,.0f}"


# ── Sheet Builders ──────────────────────────────────────────────────────────

def build_summary_sheet(wb, data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = BLUE

    prop = data.get("property", {})
    rehab = data.get("rehab_estimate", {})
    wholetail = data.get("wholetail_estimate", {})
    deal = data.get("deal_analysis", {})
    local = data.get("local_pricing", {})
    comp = data.get("comp_reference", {})

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "DATASIFT REHAB COST ESTIMATOR"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Property: {prop.get('address', 'N/A')}, {prop.get('city', '')}, {prop.get('state', '')} {prop.get('zip', '')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    pricing_note = f"Local Pricing: {local.get('city', 'N/A')}, {local.get('state', '')} (Cost Index: {local.get('cost_index', 1.0):.2f}x)"
    ws["A3"] = pricing_note
    ws["A3"].font = Font(name="Calibri", italic=True, size=10, color=BLUE)
    ws["A3"].alignment = Alignment(horizontal="center")

    # Property details
    row = 5
    ws.cell(row=row, column=1, value="Property Details").font = section_font
    row += 1
    details = [
        ("Square Footage", f"{prop.get('gla', 0):,} SF"),
        ("Bed/Bath", f"{prop.get('beds', 0)} bed / {prop.get('baths', 0)} bath"),
        ("Year Built", str(prop.get("year_built", "N/A"))),
        ("Property Type", prop.get("property_type", "N/A")),
        ("Condition", prop.get("condition_summary", "N/A")),
    ]
    for label, value in details:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=value).font = normal_font
        row += 1

    # Comp reference
    if comp.get("arv"):
        row += 1
        ws.cell(row=row, column=1, value="Comp Reference").font = section_font
        row += 1
        comp_details = [
            ("ARV (After Repair Value)", fmt_currency(comp.get("arv", 0))),
            ("Purchase Price", fmt_currency(comp.get("purchase_price", 0))),
            ("Bucket A PPSF (Unrenovated)", f"${comp.get('bucket_a_ppsf', 0):.2f}"),
            ("Bucket B PPSF (Renovated)", f"${comp.get('bucket_b_ppsf', 0):.2f}"),
            ("Renovation Premium", f"{comp.get('renovation_premium_pct', 0):.1f}%"),
            ("Market Premium ($)", fmt_currency(comp.get("market_premium_dollars", 0))),
        ]
        for label, value in comp_details:
            ws.cell(row=row, column=1, value=label).font = bold_font
            ws.cell(row=row, column=2, value=value).font = normal_font
            row += 1

    # Side-by-side comparison table
    row += 2
    ws.cell(row=row, column=1, value="Cost Estimate Comparison").font = section_font
    row += 1

    headers = ["Metric", "Full Rehab", "Wholetail"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    comparisons = [
        ("Scope", rehab.get("scope", "Full Rehab"), wholetail.get("scope", "Wholetail")),
        ("Finish Tier", rehab.get("finish_tier", ""), wholetail.get("finish_tier", "")),
        ("Estimate", rehab.get("grand_total", 0), wholetail.get("grand_total", 0)),
        ("Cost/SF", rehab.get("cost_per_sf", 0), wholetail.get("cost_per_sf", 0)),
        ("Contingency", f"{rehab.get('contingency_pct', 0.1):.0%}", f"{wholetail.get('contingency_pct', 0.05):.0%}"),
    ]

    for i, (label, rehab_val, whole_val) in enumerate(comparisons):
        is_alt = i % 2 == 1
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = bold_font

        if isinstance(rehab_val, (int, float)) and label not in ("Contingency",):
            c2 = ws.cell(row=row, column=2, value=rehab_val)
            c3 = ws.cell(row=row, column=3, value=whole_val)
            if "Cost/SF" in label:
                c2.number_format = currency_decimal_format
                c3.number_format = currency_decimal_format
            else:
                c2.number_format = currency_format
                c3.number_format = currency_format
        else:
            c2 = ws.cell(row=row, column=2, value=str(rehab_val))
            c3 = ws.cell(row=row, column=3, value=str(whole_val))

        for c in [c1, c2, c3]:
            style_data_cell(c, alt_row=is_alt)
        row += 1

    # Deal analysis — full cost breakdown
    if deal.get("arv"):
        financing = deal.get("financing", {})
        holding = deal.get("holding_costs", {})
        buying = deal.get("buying_costs", {})
        selling = deal.get("selling_costs", {})

        row += 2
        ws.cell(row=row, column=1, value="Full Deal Analysis").font = section_font
        row += 1

        headers = ["Metric", "Full Rehab", "Wholetail"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        deal_rows = [
            ("ARV (Sale Price)", deal.get("arv", 0), deal.get("arv", 0)),
            ("Purchase Price", deal.get("purchase_price", 0), deal.get("purchase_price", 0)),
            ("Rehab Cost", deal.get("rehab_cost", 0), deal.get("wholetail_cost", 0)),
            ("Financing Costs", financing.get("total_financing_cost", 0), financing.get("total_financing_cost", 0)),
            ("Holding Costs", holding.get("total_holding", 0), holding.get("total_holding", 0)),
            ("Buying Costs", buying.get("total_buying", 0), buying.get("total_buying", 0)),
            ("Selling Costs", selling.get("total_selling", 0), selling.get("total_selling", 0)),
            ("Total All-In Cost", deal.get("rehab_total_all_in", 0), deal.get("wholetail_total_all_in", 0)),
            ("Net Profit", deal.get("rehab_net_profit", 0), deal.get("wholetail_net_profit", 0)),
            ("ROI", deal.get("rehab_roi_pct", 0), deal.get("wholetail_roi_pct", 0)),
            ("75% Rule MAO", deal.get("mao_75_pct", 0), "N/A"),
        ]

        for i, (label, r_val, w_val) in enumerate(deal_rows):
            is_alt = i % 2 == 1
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = bold_font

            if label == "ROI":
                c2 = ws.cell(row=row, column=2, value=r_val)
                c2.number_format = pct_format
                c3 = ws.cell(row=row, column=3, value=w_val)
                c3.number_format = pct_format
            elif isinstance(r_val, (int, float)):
                c2 = ws.cell(row=row, column=2, value=r_val)
                c2.number_format = currency_format
                if isinstance(w_val, (int, float)):
                    c3 = ws.cell(row=row, column=3, value=w_val)
                    c3.number_format = currency_format
                else:
                    c3 = ws.cell(row=row, column=3, value=str(w_val))
            else:
                c2 = ws.cell(row=row, column=2, value=str(r_val))
                c3 = ws.cell(row=row, column=3, value=str(w_val))

            for c in [c1, c2, c3]:
                style_data_cell(c, alt_row=is_alt)
                if label == "Net Profit" and isinstance(r_val, (int, float)):
                    if r_val > 0:
                        c2.fill = green_fill
                    elif r_val < 0:
                        c2.fill = red_fill
                    if isinstance(w_val, (int, float)):
                        if w_val > 0:
                            c3.fill = green_fill
                        elif w_val < 0:
                            c3.fill = red_fill
                if label == "Total All-In Cost":
                    for c in [c1, c2, c3]:
                        c.fill = highlight_fill
                        c.font = Font(name="Calibri", bold=True, size=10)
            row += 1

    # Recommendations
    recs = data.get("recommendations", [])
    if recs:
        row += 2
        ws.cell(row=row, column=1, value="Recommendations").font = section_font
        row += 1
        for rec in recs:
            ws.cell(row=row, column=1, value=f"• {rec}").font = normal_font
            row += 1

    auto_fit_columns(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20


def build_estimate_sheet(wb, data, estimate_key, sheet_name):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = GREEN if "Rehab" in sheet_name else ORANGE

    estimate = data.get(estimate_key, {})
    prop = data.get("property", {})
    local = data.get("local_pricing", {})

    ws.merge_cells("A1:G1")
    ws["A1"] = f"DATASIFT {sheet_name.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Local Pricing: {local.get('city', 'N/A')}, {local.get('state', '')} | Multiplier: {local.get('cost_index', 1.0):.2f}x National"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color=BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Property: {prop.get('address', 'N/A')} | {prop.get('gla', 0):,} SF | Finish: {estimate.get('finish_tier', 'N/A')}"
    ws["A3"].font = Font(name="Calibri", size=10)
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:G4")
    ws["A4"] = 'Mark "X" in Include column to add item to estimate. Totals auto-calculate.'
    ws["A4"].font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws["A4"].alignment = Alignment(horizontal="center")

    # Headers — single estimate column
    row = 6
    headers = ["Repair Item", "Include?", "Qty", "Unit", "$/Unit", "Total", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    categories = estimate.get("categories", [])
    for cat in categories:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=cat.get("category", "").upper())
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="left")
        row += 1

        items = cat.get("items", [])
        for idx, item in enumerate(items):
            is_alt = idx % 2 == 1
            include = "X" if item.get("include", False) else ""

            values = [
                item.get("item", ""),
                include,
                item.get("qty", 0),
                item.get("unit", ""),
                item.get("unit_cost", 0),
                item.get("total", 0),
                item.get("notes", ""),
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                is_currency = col_idx in (5, 6)
                style_data_cell(cell, is_currency=is_currency, alt_row=is_alt)
                if col_idx == 2:
                    cell.alignment = Alignment(horizontal="center")

            row += 1

        # Category subtotal
        ws.cell(row=row, column=1, value=f"{cat.get('category', '')} Subtotal").font = bold_font
        cell = ws.cell(row=row, column=6, value=cat.get("category_total", 0))
        cell.font = bold_font
        cell.number_format = currency_format
        cell.fill = highlight_fill
        cell.border = thin_border
        row += 1
        row += 1

    # Grand totals
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="ESTIMATE TOTALS").font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    ws.cell(row=row, column=1).fill = dark_header_fill
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = dark_header_fill
    row += 1

    totals = [
        ("Subtotal", "subtotal"),
        (f"Contingency ({estimate.get('contingency_pct', 0.1):.0%})", "contingency"),
        ("GRAND TOTAL", "grand_total"),
    ]

    for label, key in totals:
        ws.cell(row=row, column=1, value=label).font = bold_font
        if "GRAND" in label:
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)

        cell = ws.cell(row=row, column=6, value=estimate.get(key, 0))
        cell.number_format = currency_format
        cell.border = thin_border
        if "GRAND" in label:
            cell.font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
            cell.fill = highlight_fill
        else:
            cell.font = bold_font
        row += 1

    # Cost per SF
    row += 1
    ws.cell(row=row, column=1, value="Cost per SF").font = bold_font
    cell = ws.cell(row=row, column=6, value=estimate.get("cost_per_sf", 0))
    cell.number_format = currency_decimal_format
    cell.font = bold_font
    cell.border = thin_border

    auto_fit_columns(ws)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["G"].width = 30


def build_condition_sheet(wb, data):
    ws = wb.create_sheet(title="Condition Assessment")
    ws.sheet_properties.tabColor = "C00000"

    ws.merge_cells("A1:F1")
    ws["A1"] = "PROPERTY CONDITION ASSESSMENT"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    prop = data.get("property", {})
    ws.merge_cells("A2:F2")
    ws["A2"] = f"{prop.get('address', 'N/A')}, {prop.get('city', '')}, {prop.get('state', '')} {prop.get('zip', '')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["Category", "Item", "Condition", "Notes", "Rehab Action", "Wholetail Action"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    conditions = data.get("condition_assessment", [])
    current_cat = None
    for idx, item in enumerate(conditions):
        cat = item.get("category", "")
        if cat != current_cat:
            current_cat = cat
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=cat.upper())
            cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
            cell.fill = dark_header_fill
            row += 1

        is_alt = idx % 2 == 1
        values = [
            "",
            item.get("item", ""),
            item.get("condition", ""),
            item.get("notes", ""),
            item.get("rehab_action", ""),
            item.get("wholetail_action", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            style_data_cell(cell, alt_row=is_alt)
            if col_idx == 3:
                cond = str(val).lower()
                if cond == "good":
                    cell.fill = green_fill
                elif cond == "fair":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif cond == "poor":
                    cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
                elif cond in ("missing", "failed", "missing/failed"):
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        row += 1

    auto_fit_columns(ws)
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25


def build_deal_analyzer_sheet(wb, data):
    ws = wb.create_sheet(title="Deal Analyzer")
    ws.sheet_properties.tabColor = GREEN

    prop = data.get("property", {})
    deal = data.get("deal_analysis", {})
    comp = data.get("comp_reference", {})
    financing = deal.get("financing", {})
    holding = deal.get("holding_costs", {})
    buying = deal.get("buying_costs", {})
    selling = deal.get("selling_costs", {})

    orange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.merge_cells("A1:F1")
    ws["A1"] = "DEAL ANALYZER"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "Full deal cost analysis including financing, holding, and transaction costs"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Property Info ──
    row = 4
    ws.cell(row=row, column=1, value="PROPERTY INFORMATION").font = section_font
    row += 1
    info = [
        ("Address", prop.get("address", "")),
        ("City/State/Zip", f"{prop.get('city', '')}, {prop.get('state', '')} {prop.get('zip', '')}"),
        ("Square Footage", f"{prop.get('gla', 0):,}"),
        ("Beds/Baths", f"{prop.get('beds', 0)} / {prop.get('baths', 0)}"),
        ("Year Built", str(prop.get("year_built", ""))),
        ("Hold Time", f"{deal.get('hold_time_months', 4)} months"),
    ]
    # Add hold time breakdown if available
    htb = deal.get("hold_time_breakdown", {})
    if htb:
        breakdown_parts = []
        if htb.get("rehab_duration_months", 0) > 0:
            breakdown_parts.append(f"Rehab: {htb['rehab_duration_months']}mo")
        if htb.get("comp_dom_days") is not None:
            breakdown_parts.append(f"DOM: {htb['comp_dom_days']}d")
        if htb.get("marketing_period_months", 0) > 0:
            breakdown_parts.append(f"Marketing: {htb['marketing_period_months']}mo")
        if htb.get("closing_period_months", 0) > 0:
            breakdown_parts.append(f"Closing: {htb['closing_period_months']}mo")
        if breakdown_parts:
            info.append(("Hold Time Breakdown", " + ".join(breakdown_parts)))
        if htb.get("source"):
            info.append(("Hold Time Source", htb["source"].replace("_", " ").title()))
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=val).font = normal_font
        row += 1

    # ── Property Valuation ──
    row += 1
    ws.cell(row=row, column=1, value="PROPERTY VALUATION").font = section_font
    row += 1
    val_rows = [
        ("After Repair Value (ARV)", deal.get("arv", 0)),
        ("Purchase Price", deal.get("purchase_price", 0)),
        ("Estimated Repair Costs (Full Rehab)", deal.get("rehab_cost", 0)),
        ("Estimated Repair Costs (Wholetail)", deal.get("wholetail_cost", 0)),
    ]
    for label, val in val_rows:
        ws.cell(row=row, column=1, value=label).font = bold_font
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = currency_format
        c.font = normal_font
        c.border = thin_border
        row += 1

    # ── Financing Costs ──
    row += 1
    ws.cell(row=row, column=1, value="FINANCING COSTS").font = section_font
    row += 1
    fin_rows = [
        ("Loan Amount (Purchase + Rehab)", financing.get("loan_amount", 0)),
        ("LTV", f"{financing.get('ltv_pct', 1.0):.0%}"),
        ("Interest Rate (Annual)", f"{financing.get('interest_rate_annual', 0.12):.1%}"),
        ("Points", f"{financing.get('points_pct', 0):.1%}"),
        ("Monthly Interest", financing.get("monthly_interest", 0)),
        ("Total Financing Cost", financing.get("total_financing_cost", 0)),
    ]
    for label, val in fin_rows:
        is_alt = fin_rows.index((label, val)) % 2 == 1
        ws.cell(row=row, column=1, value=label).font = bold_font
        if isinstance(val, (int, float)):
            c = ws.cell(row=row, column=2, value=val)
            c.number_format = currency_format
            c.font = normal_font
        else:
            c = ws.cell(row=row, column=2, value=val)
            c.font = normal_font
        c.border = thin_border
        if is_alt:
            c.fill = alt_fill
        row += 1
    ws.cell(row=row-1, column=1).font = Font(name="Calibri", bold=True, size=10, color=BLUE)

    # ── Holding Costs ──
    row += 1
    ws.cell(row=row, column=1, value="HOLDING COSTS (MONTHLY)").font = section_font
    row += 1
    hold_rows = [
        ("Property Taxes", holding.get("property_taxes_monthly", 0)),
        ("Insurance", holding.get("insurance_monthly", 0)),
        ("HOA/Condo Fees", holding.get("hoa_monthly", 0)),
        ("Gas", holding.get("gas_monthly", 0)),
        ("Water", holding.get("water_monthly", 0)),
        ("Electricity", holding.get("electric_monthly", 0)),
        ("Other", holding.get("other_monthly", 0)),
        ("Total Monthly Holding", holding.get("total_monthly", 0)),
    ]
    for idx, (label, val) in enumerate(hold_rows):
        is_alt = idx % 2 == 1
        ws.cell(row=row, column=1, value=label).font = bold_font
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = currency_format
        c.font = normal_font
        c.border = thin_border
        if is_alt:
            c.fill = alt_fill
        row += 1
    # Total holding for full hold period
    ws.cell(row=row, column=1, value=f"Total Holding ({deal.get('hold_time_months', 4)} months)").font = Font(name="Calibri", bold=True, size=10, color=BLUE)
    c = ws.cell(row=row, column=2, value=holding.get("total_holding", 0))
    c.number_format = currency_format
    c.font = Font(name="Calibri", bold=True, size=10, color=BLUE)
    c.border = thin_border
    row += 1

    # ── Buying Transaction Costs ──
    row += 1
    ws.cell(row=row, column=1, value="BUYING TRANSACTION COSTS").font = section_font
    row += 1
    buy_rows = [
        ("Escrow/Attorney Fees", buying.get("escrow_attorney", 0)),
        ("Title Insurance", buying.get("title_insurance", 0)),
        ("Total Buying Costs", buying.get("total_buying", 0)),
    ]
    for idx, (label, val) in enumerate(buy_rows):
        ws.cell(row=row, column=1, value=label).font = bold_font
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = currency_format
        c.font = normal_font
        c.border = thin_border
        if idx % 2 == 1:
            c.fill = alt_fill
        row += 1
    ws.cell(row=row-1, column=1).font = Font(name="Calibri", bold=True, size=10, color=BLUE)

    # ── Selling Transaction Costs ──
    row += 1
    ws.cell(row=row, column=1, value="SELLING TRANSACTION COSTS").font = section_font
    row += 1
    sell_rows = [
        ("Realtor Fees", selling.get("realtor_fees", 0)),
        ("Title Insurance", selling.get("title_insurance", 0)),
        ("Escrow/Attorney Fees", selling.get("escrow_attorney", 0)),
        ("Recording Fees", selling.get("recording_fees", 0)),
        ("Transfer/Conveyance Tax", selling.get("transfer_tax", 0)),
        ("Home Warranty", selling.get("home_warranty", 0)),
        ("Total Selling Costs", selling.get("total_selling", 0)),
    ]
    for idx, (label, val) in enumerate(sell_rows):
        ws.cell(row=row, column=1, value=label).font = bold_font
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = currency_format
        c.font = normal_font
        c.border = thin_border
        if idx % 2 == 1:
            c.fill = alt_fill
        row += 1
    ws.cell(row=row-1, column=1).font = Font(name="Calibri", bold=True, size=10, color=BLUE)

    # ── Profit & Loss Summary ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value="PROFIT & LOSS SUMMARY").font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    ws.cell(row=row, column=1).fill = dark_header_fill
    for c in range(1, 4):
        ws.cell(row=row, column=c).fill = dark_header_fill
    row += 1

    # Headers
    headers_pl = ["", "Full Rehab", "Wholetail"]
    for i, h in enumerate(headers_pl, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 3)
    row += 1

    # Calculate wholetail financing (different loan amount)
    wholetail_financing = 0
    if financing.get("total_financing_cost", 0) > 0 and deal.get("rehab_cost", 0) > 0:
        rehab_loan = financing.get("loan_amount", 0)
        wholetail_loan = deal.get("purchase_price", 0) + deal.get("wholetail_cost", 0)
        wholetail_financing = financing.get("total_financing_cost", 0) * (wholetail_loan / rehab_loan) if rehab_loan > 0 else 0

    pl_rows = [
        ("ARV (Sale Price)", deal.get("arv", 0), deal.get("arv", 0), False),
        ("Purchase Price", deal.get("purchase_price", 0), deal.get("purchase_price", 0), False),
        ("Rehab Cost", deal.get("rehab_cost", 0), deal.get("wholetail_cost", 0), False),
        ("Financing Costs", financing.get("total_financing_cost", 0), wholetail_financing, False),
        ("Holding Costs", holding.get("total_holding", 0), holding.get("total_holding", 0), False),
        ("Buying Costs", buying.get("total_buying", 0), buying.get("total_buying", 0), False),
        ("Selling Costs", selling.get("total_selling", 0), selling.get("total_selling", 0), False),
        ("", "", "", False),
        ("Total All-In Cost", deal.get("rehab_total_all_in", 0), deal.get("wholetail_total_all_in", 0), True),
        ("NET PROFIT", deal.get("rehab_net_profit", 0), deal.get("wholetail_net_profit", 0), True),
        ("ROI", deal.get("rehab_roi_pct", 0), deal.get("wholetail_roi_pct", 0), True),
    ]

    for idx, (label, r_val, w_val, is_total) in enumerate(pl_rows):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name="Calibri", bold=True, size=11 if is_total else 10)

        if label == "ROI":
            c2 = ws.cell(row=row, column=2, value=r_val)
            c2.number_format = pct_format
            c3 = ws.cell(row=row, column=3, value=w_val)
            c3.number_format = pct_format
            c2.font = Font(name="Calibri", bold=True, size=11)
            c3.font = Font(name="Calibri", bold=True, size=11)
        elif isinstance(r_val, (int, float)) and label:
            c2 = ws.cell(row=row, column=2, value=r_val)
            c2.number_format = currency_format
            c3 = ws.cell(row=row, column=3, value=w_val)
            c3.number_format = currency_format
            if is_total:
                c2.font = Font(name="Calibri", bold=True, size=11)
                c3.font = Font(name="Calibri", bold=True, size=11)
        else:
            c2 = ws.cell(row=row, column=2, value="")
            c3 = ws.cell(row=row, column=3, value="")

        for c in [c1, c2, c3]:
            c.border = thin_border

        if label == "NET PROFIT":
            for c in [c2, c3]:
                val = c.value
                if isinstance(val, (int, float)):
                    c.fill = green_fill if val > 0 else red_fill
                    c.font = Font(name="Calibri", bold=True, size=12, color=GREEN if val > 0 else RED)
        elif is_total and label != "ROI":
            for c in [c1, c2, c3]:
                c.fill = highlight_fill

        row += 1

    # 75% Rule
    row += 1
    ws.cell(row=row, column=1, value="75% RULE CHECK").font = section_font
    row += 1
    mao = deal.get("mao_75_pct", 0)
    ws.cell(row=row, column=1, value="Max Allowable Offer (75% Rule)").font = bold_font
    c = ws.cell(row=row, column=2, value=mao)
    c.number_format = currency_format
    c.font = Font(name="Calibri", bold=True, size=12, color=GREEN if mao > deal.get("purchase_price", 0) else RED)

    row += 1
    ws.cell(row=row, column=1, value="Formula: ARV × 75% − Rehab Cost").font = Font(name="Calibri", italic=True, size=9, color="666666")

    if comp.get("market_premium_dollars"):
        row += 2
        ws.cell(row=row, column=1, value="REHAB BUDGET vs MARKET PREMIUM").font = section_font
        row += 1
        ws.cell(row=row, column=1, value="Market Premium (Bucket B − A × GLA)").font = bold_font
        c = ws.cell(row=row, column=2, value=comp.get("market_premium_dollars", 0))
        c.number_format = currency_format
        row += 1
        ws.cell(row=row, column=1, value="Rehab as % of Premium").font = bold_font
        ratio = deal.get("rehab_margin_vs_premium", 0)
        c = ws.cell(row=row, column=2, value=ratio)
        c.number_format = pct_format
        if isinstance(ratio, (int, float)):
            c.fill = green_fill if ratio <= 0.7 else red_fill
        row += 1
        ws.cell(row=row, column=1, value="Target: 40-70% (healthy margin)").font = Font(name="Calibri", italic=True, size=9, color="666666")

    # Financing assumptions note
    row += 2
    ws.cell(row=row, column=1, value="ASSUMPTIONS").font = section_font
    row += 1
    # Build hold time assumption string with breakdown if available
    hold_str = f"Hold Time: {deal.get('hold_time_months', 4)} months"
    htb = deal.get("hold_time_breakdown", {})
    if htb and htb.get("source") == "calculated":
        parts = []
        if htb.get("rehab_duration_months", 0) > 0:
            parts.append(f"{htb['rehab_duration_months']}mo rehab")
        if htb.get("comp_dom_days") is not None:
            parts.append(f"{htb['comp_dom_days']}d DOM")
        if htb.get("marketing_period_months", 0) > 0:
            parts.append(f"{htb['marketing_period_months']}mo marketing")
        if htb.get("closing_period_months", 0) > 0:
            parts.append(f"{htb['closing_period_months']}mo closing")
        if parts:
            hold_str += f" ({' + '.join(parts)})"
    assumptions = [
        f"Financing: {financing.get('notes', 'Hard money, 100% LTV')}",
        f"Interest Rate: {financing.get('interest_rate_annual', 0.12):.1%} annual",
        hold_str,
        f"Realtor Commission: {selling.get('realtor_pct', 0.05):.0%} of ARV",
        f"Title Insurance: {buying.get('title_insurance_pct', 0.0045):.2%} of price",
    ]
    for note in assumptions:
        ws.cell(row=row, column=1, value=note).font = Font(name="Calibri", italic=True, size=9, color="666666")
        row += 1

    auto_fit_columns(ws)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18


def build_budget_tracker_sheet(wb, data):
    ws = wb.create_sheet(title="Budget Tracker")
    ws.sheet_properties.tabColor = ORANGE

    ws.merge_cells("A1:H1")
    ws["A1"] = "BUDGET TRACKER"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "Track actual costs against budget during renovation"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["#", "Line Item", "Budget", "Actual", "Variance", "Invoice Date", "Paid Date", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    rehab = data.get("rehab_estimate", {})
    categories = rehab.get("categories", [])

    line_num = 1
    for cat in categories:
        cat_total = cat.get("category_total", 0)
        if cat_total > 0:
            ws.cell(row=row, column=1, value=line_num).font = normal_font
            ws.cell(row=row, column=2, value=cat.get("category", "")).font = normal_font
            c = ws.cell(row=row, column=3, value=cat_total)
            c.number_format = currency_format
            c.font = normal_font
            ws.cell(row=row, column=5, value=f"=C{row}-D{row}").font = normal_font
            ws.cell(row=row, column=5).number_format = currency_format
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
            line_num += 1
            row += 1

    ws.cell(row=row, column=1, value=line_num).font = normal_font
    ws.cell(row=row, column=2, value="Contingency").font = normal_font
    c = ws.cell(row=row, column=3, value=rehab.get("contingency", 0))
    c.number_format = currency_format
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    row += 1
    ws.cell(row=row, column=2, value="TOTALS").font = bold_font
    start_data = 5
    end_data = row - 2
    for col in [3, 4, 5]:
        letter = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({letter}{start_data}:{letter}{end_data})")
        c.font = bold_font
        c.number_format = currency_format
        c.fill = highlight_fill
        c.border = thin_border

    auto_fit_columns(ws)
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["H"].width = 30


def build_material_specs_sheet(wb, data):
    ws = wb.create_sheet(title="Material Specs")
    ws.sheet_properties.tabColor = BLUE

    ws.merge_cells("A1:G1")
    ws["A1"] = "RECOMMENDED MATERIAL SPECIFICATIONS"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    rehab = data.get("rehab_estimate", {})
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Finish Tier: {rehab.get('finish_tier', 'N/A')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    headers = ["Item", "Specification", "Est. Price", "Finish", "SKU/Link", "Alt. Option", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    specs = data.get("material_specs", [])
    if specs:
        current_cat = None
        for idx, spec in enumerate(specs):
            cat = spec.get("category", "")
            if cat != current_cat:
                current_cat = cat
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                cell = ws.cell(row=row, column=1, value=cat.upper())
                cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
                cell.fill = dark_header_fill
                row += 1
            is_alt = idx % 2 == 1
            values = [spec.get("item", ""), spec.get("specification", ""), spec.get("est_price", ""),
                      spec.get("finish", ""), spec.get("sku", ""), spec.get("alt_option", ""), spec.get("notes", "")]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                style_data_cell(cell, alt_row=is_alt)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Material specs to be populated based on finish tier selection").font = Font(name="Calibri", italic=True, size=10, color="666666")

    row += 2
    ws.cell(row=row, column=1, value="PAINT COLOR SCHEME (Modern Neutral — Most Popular)").font = section_font
    row += 1
    colors = [
        ("Interior Walls", "Agreeable Gray", "Sherwin-Williams", "SW 7029"),
        ("Trim & Doors", "Extra White", "Sherwin-Williams", "SW 7006"),
        ("Ceilings", "Ceiling Bright White", "Sherwin-Williams", "SW 7007"),
        ("Cabinets", "Pure White", "Sherwin-Williams", "SW 7005"),
        ("Accent", "Iron Ore", "Sherwin-Williams", "SW 7069"),
        ("Exterior", "Repose Gray", "Sherwin-Williams", "SW 7015"),
    ]
    sub_headers = ["Application", "Color Name", "Brand", "Code"]
    for i, h in enumerate(sub_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(sub_headers))
    row += 1
    for idx, (app, name, brand, code) in enumerate(colors):
        is_alt = idx % 2 == 1
        for col_idx, val in enumerate([app, name, brand, code], 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            style_data_cell(cell, alt_row=is_alt)
        row += 1

    auto_fit_columns(ws)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["E"].width = 25


def build_checklist_sheet(wb, data):
    ws = wb.create_sheet(title="Project Checklist")
    ws.sheet_properties.tabColor = "548235"

    ws.merge_cells("A1:F1")
    ws["A1"] = "PROJECT CHECKLIST & TIMELINE"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    headers = ["Task", "Status", "Due Date", "Completed", "Assigned To", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    phases = [
        ("PRE-CONSTRUCTION", [
            "Turn on utilities (water, electric, gas)", "Property inspection / scope walkthrough",
            "Finalize scope of work", "Get contractor bids (minimum 2-3)",
            "Pull required permits", "Order materials (long lead items first)",
            "Schedule dumpster delivery", "Secure property (locks, lockbox)",
        ]),
        ("WEEK 1 — DEMO & ROUGH", [
            "Dumpster delivered", "Interior demo (kitchen, bath, flooring)",
            "Rough plumbing", "Rough electrical", "HVAC rough-in / replacement",
            "Framing changes (if any)", "Schedule rough inspections",
        ]),
        ("WEEK 2 — SYSTEMS & STRUCTURE", [
            "Rough inspections passed", "Insulation", "Drywall hang + tape + mud",
            "Window installation (if replacing)", "Exterior repairs (siding, soffit, fascia)",
            "Roof work (if needed)",
        ]),
        ("WEEK 3 — FINISHES", [
            "Drywall sanding + prime", "Paint — primer coat",
            "Paint — finish coats (walls + ceilings)", "Flooring installation",
            "Cabinet installation", "Tile work (bathrooms, backsplash)",
        ]),
        ("WEEK 4 — FIXTURES & DETAILS", [
            "Countertop installation", "Plumbing fixtures (faucets, toilets, sinks)",
            "Electrical fixtures (lights, outlets, switches)", "Appliance installation",
            "Door hardware", "Backsplash completion", "Paint touch-up (all trim, doors)",
        ]),
        ("WEEK 5 — PUNCH LIST & CLOSE OUT", [
            "Final walkthrough with contractor", "Punch list items addressed",
            "Final inspections (building, mechanical)", "Carpet installation (if applicable)",
            "Deep cleaning (professional)", "Landscaping / curb appeal",
        ]),
        ("LISTING PREPARATION", [
            "Staging (furniture + decor)", "Professional photography",
            "Listing prep (description, disclosures)", "Schedule open house",
            "MLS listing live", "Lockbox for showings",
        ]),
    ]

    for phase_name, tasks in phases:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=phase_name)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill = dark_header_fill
        row += 1
        for idx, task in enumerate(tasks):
            is_alt = idx % 2 == 1
            ws.cell(row=row, column=1, value=task).font = normal_font
            ws.cell(row=row, column=2, value="Not Started").font = normal_font
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if is_alt:
                    cell.fill = alt_fill
            row += 1
        row += 1

    auto_fit_columns(ws)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["F"].width = 30


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_rehab_excel.py <output_path> <json_data_path>")
        sys.exit(1)

    output_path = sys.argv[1]
    json_path = sys.argv[2]

    with open(json_path, "r") as f:
        data = json.load(f)

    wb = Workbook()

    build_summary_sheet(wb, data)
    build_estimate_sheet(wb, data, "rehab_estimate", "Full Rehab Estimate")
    build_estimate_sheet(wb, data, "wholetail_estimate", "Wholetail Estimate")
    build_condition_sheet(wb, data)
    build_deal_analyzer_sheet(wb, data)
    build_budget_tracker_sheet(wb, data)
    build_material_specs_sheet(wb, data)
    build_checklist_sheet(wb, data)

    wb.save(output_path)
    print(f"Excel workbook saved to: {output_path}")


if __name__ == "__main__":
    main()
