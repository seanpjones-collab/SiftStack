"""Generate Ohio first-to-market distress data sources XLSX report.

Produces output/ohio_distress_sources_{YYYYMMDD}.xlsx with 3 sheets
(Summit, Cuyahoga, Stark) matching the first-market-county-data skill's
output contract: Data Type, Priority, Office, Address, Phone, Portal,
FOIA Email, Difficulty, Notes.

Data compiled from parallel research runs 2026-04-20.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent / "output"
OUT_DIR.mkdir(exist_ok=True)

COLUMNS = [
    ("Data Type", 22),
    ("Priority", 9),
    ("Office Name", 38),
    ("Address", 40),
    ("Phone", 16),
    ("Portal URL", 55),
    ("FOIA Email", 30),
    ("Difficulty", 11),
    ("Notes", 90),
]

# ── Per-county data (compiled from first-market-county-data research 2026-04-20) ──

SUMMIT = [
    {
        "Data Type": "Foreclosure (Lis Pendens)",
        "Priority": "A",
        "Office Name": "Summit County Clerk of Courts — Common Pleas General Division",
        "Address": "205 S. High St, Akron OH 44308",
        "Phone": "330-643-2200",
        "Portal URL": "https://clerkweb.summitoh.net/",
        "FOIA Email": "",
        "Difficulty": "High",
        "Notes": "Classic ASP, POST form with session. No clean case-type + date-range filter — must search by name or iterate case numbers. No CAPTCHA. Playwright required. 2002+ online. Validate via Akron Legal News as cross-check.",
    },
    {
        "Data Type": "Probate (Auth. to Administer / Unknown Heirs)",
        "Priority": "A",
        "Office Name": "Summit County Probate Court",
        "Address": "209 S. High St, Room 215, Akron OH 44308",
        "Phone": "330-643-2350",
        "Portal URL": "https://search.summitohioprobate.com/eservices/home.page.2",
        "FOIA Email": "",
        "Difficulty": "Medium",
        "Notes": "CourtView CMS (JS-driven, ASPX session, meta-refresh redirect). Supports case-type + date-range smart search. Filter case type = Estate–Full / Estate–Release / Authority to Administer. No CAPTCHA. Playwright required for JS redirect.",
    },
    {
        "Data Type": "Sheriff Sales (Foreclosure + Tax)",
        "Priority": "A",
        "Office Name": "Summit County Sheriff — Civil Division (auction on RealAuction)",
        "Address": "53 University Ave, Akron OH 44308",
        "Phone": "330-643-2278",
        "Portal URL": "https://summit.sheriffsaleauction.ohio.gov + https://sheriff.summitoh.net/pages/Sheriff-Sales.html",
        "FOIA Email": "",
        "Difficulty": "Low",
        "Notes": "Mortgage sales Fri 10am, delinquent-tax sales Tue 10am. Public catalog open (no login to view). Bidding requires registered account + deposit. CLEANEST SCRAPE TARGET for Summit. No CAPTCHA, no auth.",
    },
    {
        "Data Type": "Tax Delinquency",
        "Priority": "A",
        "Office Name": "Summit County Fiscal Office — Delinquent Tax Dept",
        "Address": "175 S. Main St, Akron OH 44308",
        "Phone": "330-643-2600",
        "Portal URL": "http://fiscaloffice.summitoh.net/index.php/property-tax-search + https://www.akronlegalnews.com/notices/delinquent_taxes_payment",
        "FOIA Email": "",
        "Difficulty": "Medium",
        "Notes": "Per-parcel search only on Fiscal site — no bulk export. ALN republishes annual certified delinquent list as browsable HTML. No daily delta feed. Annual publication after 2nd-half tax settlement for parcels > $5.",
    },
    {
        "Data Type": "Code Violations",
        "Priority": "B",
        "Office Name": "City of Akron — Housing Compliance Division",
        "Address": "166 S. High St, Akron OH 44308",
        "Phone": "330-375-2366 (3-1-1)",
        "Portal URL": "None — intake only",
        "FOIA Email": "HousingCodeComplaints@akronohio.gov",
        "Difficulty": "Very High",
        "Notes": "NO public data portal. Requires Ohio Public Records Request (ORC 149.43) per batch. Intake via 3-1-1 or fillable PDF. NOT scrapable — skip for daily automation. Condemnation lists sometimes in City Council packets.",
    },
    {
        "Data Type": "Evictions (FED)",
        "Priority": "B",
        "Office Name": "Akron Muni + Barberton Muni + Stow Muni Courts",
        "Address": "Akron: 172 S. Broadway. Barberton: 576 W. Park Ave. Stow: 4400 Courthouse Dr.",
        "Phone": "Akron 330-375-2120 / Barberton 330-753-2261 / Stow 330-564-4465",
        "Portal URL": "Akron: https://portal-ohakron.tylertech.cloud/Portal | Barberton: https://caselook.barbertonclerkofcourt.com | Stow: http://eservices.stowmunicourt.com/eservices",
        "FOIA Email": "",
        "Difficulty": "High (Akron) / Medium (Barberton, Stow)",
        "Notes": "Three separate scrapers required. Akron Muni uses Tyler Odyssey Portal (registration required for case lists, rate-limited). Barberton uses Henschen CaseLook (disclaimer gate). Stow uses eServices (CourtView-like). Case type CVG/CVH = Forcible Entry & Detainer. 24-hr posting delay.",
    },
]

CUYAHOGA = [
    {
        "Data Type": "Foreclosure (Lis Pendens)",
        "Priority": "A",
        "Office Name": "Cuyahoga County Clerk of Courts — General Division Civil",
        "Address": "1200 Ontario St, Cleveland OH 44113",
        "Phone": "216-443-7950",
        "Portal URL": "https://cpdocket.cp.cuyahogacounty.gov/Search.aspx",
        "FOIA Email": "",
        "Difficulty": "High",
        "Notes": "ASP.NET WebForms POST (ViewState + __doPostBack). *** ToS EXPLICITLY PROHIBITS bulk DB access / automated scraping *** — legal/ethical risk. No CAPTCHA. Supplement with DLN legal notices for safer Day-0 coverage. Consider skipping automated foreclosure scrape for Cuyahoga.",
    },
    {
        "Data Type": "Probate (Auth. to Administer / Unknown Heirs)",
        "Priority": "A",
        "Office Name": "Cuyahoga County Probate Court — Case Records Search",
        "Address": "1 Lakeside Ave W, Cleveland OH 44113",
        "Phone": "216-443-8764",
        "Portal URL": "https://probate.cuyahogacounty.gov/pa/",
        "FOIA Email": "pccpc@cuyahogacounty.us",
        "Difficulty": "Medium",
        "Notes": "Separate ASP.NET system from main docket (different ToS). No login/CAPTCHA on free search. Filing-date search works for daily new estates. PR name + decedent name on first docket entry. SIMPLEST OF THE 3 OHIO PROBATE PORTALS.",
    },
    {
        "Data Type": "Sheriff Sales (Foreclosure + Tax Delinquent)",
        "Priority": "A",
        "Office Name": "Cuyahoga County Sheriff — Civil Division",
        "Address": "1215 W 3rd St, Cleveland OH 44113",
        "Phone": "216-443-6000",
        "Portal URL": "https://cpdocket.cp.cuyahogacounty.gov/SheriffSearch/results.aspx (+ printresults.aspx single-page dump)",
        "FOIA Email": "",
        "Difficulty": "Low",
        "Notes": "*** BEST SCRAPE TARGET IN ALL 3 COUNTIES *** Query-string URL (searchType, dateFrom, dateTo, foreclosureType), 126+ paginated pages, pagination uses __doPostBack. Status field flags WITHDRAWN-TAX DELINQUENT / SOLD / CANCELLED. No auth, no CAPTCHA. printresults.aspx dumps all on one page. RealAuction mirror (cuyahoga.sheriffsaleauction.ohio.gov) returns 403.",
    },
    {
        "Data Type": "Tax Delinquency",
        "Priority": "A",
        "Office Name": "Cuyahoga County Fiscal Officer — Real Property / Treasurer",
        "Address": "2079 E 9th St, Cleveland OH 44115",
        "Phone": "216-443-7420",
        "Portal URL": "https://cuyahogacounty.gov/fiscal-officer/departments/real-property/delinquent-publication + https://cuyahogacounty.gov/treasury/delinquency",
        "FOIA Email": "",
        "Difficulty": "High",
        "Notes": "Annual PDF publication only (ORC 5721.03, once/year in Daily Legal News). No CSV/API. Tax-lien certificate sale is bulk negotiated ($10-17M, thousands of parcels) — not parcel-by-parcel auction. NOT a daily-scrape target. Probe https://fiscalhub.gis.cuyahogacounty.gov/pages/2026-tax for ArcGIS REST layer.",
    },
    {
        "Data Type": "Code Violations",
        "Priority": "B",
        "Office Name": "City of Cleveland Dept of Building & Housing — Code Enforcement",
        "Address": "601 Lakeside Ave Rm 517, Cleveland OH 44114",
        "Phone": "216-664-2000 / 311",
        "Portal URL": "https://data.clevelandohio.gov/datasets/building-complaint-violation-notices + https://aca-prod.accela.com/COC/Default.aspx",
        "FOIA Email": "",
        "Difficulty": "Low",
        "Notes": "*** UNIQUELY VALUABLE *** Cleveland Open Data portal exposes Socrata/ArcGIS REST/GeoJSON/CSV API — nightly bulk pulls possible. Accela Citizen Access handles per-address enrichment. $20/letter for formal violation letter. ONLY REAL CODE VIOLATION DATA IN ALL 3 COUNTIES.",
    },
    {
        "Data Type": "Evictions (FED)",
        "Priority": "B",
        "Office Name": "Cleveland Housing Court (specialized division, not muni)",
        "Address": "Justice Center, 1200 Ontario St, 13th Floor, Cleveland OH 44113",
        "Phone": "216-664-4295",
        "Portal URL": "https://www.clevelandhousingcourt.org/docket",
        "FOIA Email": "",
        "Difficulty": "Medium",
        "Notes": "HTML dockets (Civil / Criminal / Corporate). Simple requests + BeautifulSoup — no JS, no CAPTCHA. CAVEAT: published docket = cases scheduled for hearing, not new filings. For Day-0 filing search, fall back to cpdocket CV case type. Covers City of Cleveland ONLY — suburbs go to Bedford/Parma/South Euclid/etc. muni courts.",
    },
]

STARK = [
    {
        "Data Type": "Foreclosure (Lis Pendens)",
        "Priority": "A",
        "Office Name": "Stark County Clerk of Courts — Common Pleas Civil Division",
        "Address": "115 Central Plaza N Ste 101, Canton OH 44702",
        "Phone": "330-451-7801",
        "Portal URL": "https://www.starkcjis.org/ (unified CJIS, covers Common Pleas + 3 muni courts) — https://www.starkcountycjis.org/Common_Pleas/docket/",
        "FOIA Email": "",
        "Difficulty": "High",
        "Notes": "*** SERVER RETURNS 403 TO BOT USER AGENTS *** Requires headed Playwright with realistic headers. Guest login required (free, no CAPTCHA past UA check). Frameset + POST form. Filter civil case types to CV + Real Estate sub-code. Single scraper also covers evictions (row #6).",
    },
    {
        "Data Type": "Probate (Auth. to Administer / Unknown Heirs)",
        "Priority": "A",
        "Office Name": "Stark County Probate Court",
        "Address": "110 Central Plaza S, Canton OH 44702",
        "Phone": "330-451-7755 (Records 330-451-7753)",
        "Portal URL": "http://www.probate.co.stark.oh.us/search/search.html",
        "FOIA Email": "",
        "Difficulty": "Medium",
        "Notes": "Legacy HTML frameset, GET form to case-index CGI. PORT 80 ONLY (http, not https) — separate codebase from CJIS. Case-type filter exposes Estate/Full and Estate/Release. No CAPTCHA. PDF-linked filings. May need real-browser UA to bypass geo/UA block.",
    },
    {
        "Data Type": "Sheriff Sales (Foreclosure + Tax + Tax Certificate)",
        "Priority": "A",
        "Office Name": "Stark County Sheriff — Civil/Sales",
        "Address": "4500 Atlantic Blvd NE, Canton OH 44705",
        "Phone": "",
        "Portal URL": "https://sheriffsales.starkcountyohio.gov/SearchSales.aspx",
        "FOIA Email": "STARKSALES@starkcountyohio.gov",
        "Difficulty": "Low",
        "Notes": "*** CLEANEST SCRAPE TARGET FOR STARK *** ASP.NET WebForms, __VIEWSTATE/__doPostBack. Sale Type filter (Foreclosure / Delinquent Tax / Tax Certificate). Lists case #, address, parcel, city, zip, defendant, sale type. No CAPTCHA, no login. Version 3.2.11. stark.sheriffsaleauction.ohio.gov (RealAuction) returns 403 — use .starkcountyohio.gov variant.",
    },
    {
        "Data Type": "Tax Delinquency",
        "Priority": "A",
        "Office Name": "Stark County Auditor (publishes) / Treasurer (collects)",
        "Address": "110 Central Plaza S Ste 220, Canton OH 44702",
        "Phone": "Treasurer 330-451-7814",
        "Portal URL": "https://www.starkcountyohio.gov/_T26_R577.php (annual PDF)",
        "FOIA Email": "",
        "Difficulty": "Medium",
        "Notes": "Annual PDF only. Published Nov in Canton Repository + Auditor site. NOT a daily signal. Scrape once per year, diff YoY. Requires text extraction (pypdfium2 + OCR fallback). 1-year redemption window before certificate sale. Separate Mobile Homes PDF.",
    },
    {
        "Data Type": "Code Violations",
        "Priority": "B",
        "Office Name": "Fragmented — Canton / North Canton / Massillon / Alliance (all separate)",
        "Address": "Canton City Hall: 218 Cleveland Ave SW, Canton OH 44702",
        "Phone": "Canton 330-489-3283",
        "Portal URL": "None — Canton info only: https://cantonohio.gov/163/Code-Enforcement. Intake via SeeClickFix.",
        "FOIA Email": "",
        "Difficulty": "Very High",
        "Notes": "Zero cities in Stark publish a searchable violations DB. Phone, email, or Public Records Request only. NOT scrapable — skip for daily automation. Consider SeeClickFix public API if exposed (unconfirmed). Revisit if Canton migrates to GovPilot/OpenGov.",
    },
    {
        "Data Type": "Evictions (FED)",
        "Priority": "B",
        "Office Name": "Canton Muni + Massillon Muni + Alliance Muni (one shared CJIS docket)",
        "Address": "Canton: 218 Cleveland Ave SW. Massillon: 125 Lincoln Way E. Alliance: 470 E Broadway.",
        "Phone": "Massillon 330-830-1731 / Alliance 330-823-6600",
        "Portal URL": "Unified: https://www.starkcjis.org/ — Canton: /canton/docket/ — Massillon: /Massillon/docket/search_large_frame.html — Alliance: /Alliance/docket/",
        "FOIA Email": "",
        "Difficulty": "High",
        "Notes": "Same 403 bot-block as row #1. Unified CJIS search covers all 3 munis in ONE query — scrape once, split by court. Case-type filter = Civil + FED. Plaintiff = landlord (target contact), defendant = tenant. Shares scraper with foreclosure docket.",
    },
]

SUMMARY = [
    ("", "", "TL;DR"),
    ("", "", "Cuyahoga is unexpectedly the EASIEST county to automate: has the best sheriff sale portal AND a true Open Data API for code violations."),
    ("", "", "Summit has 4 different platforms (Tyler Odyssey, CourtView, classic ASP, ALN newspaper) — most fragmented."),
    ("", "", "Stark needs HEADED Playwright with real UA to bypass 403 bot-block on starkcjis.org."),
    ("", "", ""),
    ("", "", "Consistent patterns across all 3 counties:"),
    ("", "", "• Sheriff sale portals = easiest scrape, but LATE signal (auction weeks away, months/years after filing)"),
    ("", "", "• Court dockets = earliest data, hardest scrape (all ASP.NET, all need Playwright)"),
    ("", "", "• Probate = best balance of early-stage data + achievable scrape (MEDIUM across all 3)"),
    ("", "", "• Code violations = only Cuyahoga has a real data source (Cleveland Open Data API)"),
    ("", "", "• Tax delinquency = annual PDFs everywhere, not a daily signal"),
    ("", "", ""),
    ("", "", "Recommended build order:"),
    ("", "", "1. Cuyahoga SheriffSearch (LOW, printresults.aspx dumps 126 pages to one URL)"),
    ("", "", "2. Cleveland Open Data code violations (LOW, REST API)"),
    ("", "", "3. Cuyahoga Probate /pa/ (MEDIUM, simplest of 3 probate portals)"),
    ("", "", "4. Summit Probate CourtView (MEDIUM, reuse patterns from #3)"),
    ("", "", "5. Stark Probate (MEDIUM, legacy port-80)"),
    ("", "", "6. Summit & Stark Sheriff portals (LOW)"),
    ("", "", "7. Summit Common Pleas (HIGH, but no ToS concern)"),
    ("", "", "8. Stark CJIS (HIGH, needs headed Chromium — covers foreclosures + evictions together)"),
    ("", "", "9. Defer/skip: Cuyahoga Common Pleas (ToS risk), all code violations except Cleveland, tax delinquency (annual cron)"),
]


def write_sheet(wb: Workbook, title: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Header row
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Data rows
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    difficulty_colors = {
        "Low": "C6EFCE",
        "Medium": "FFEB9C",
        "High": "FFC7CE",
        "Very High": "FFA6A6",
    }

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (name, _) in enumerate(COLUMNS, 1):
            value = row.get(name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_align
            if name == "Difficulty":
                key = next((k for k in difficulty_colors if value.startswith(k)), None)
                if key:
                    cell.fill = PatternFill(
                        start_color=difficulty_colors[key],
                        end_color=difficulty_colors[key],
                        fill_type="solid",
                    )
                cell.font = Font(bold=True)
        ws.row_dimensions[row_idx].height = 90

    ws.freeze_panes = "A2"


def write_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary", 0)  # first sheet
    ws["A1"] = f"Ohio First-to-Market Distress Data Sources — Generated {datetime.now():%Y-%m-%d}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 26

    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, (_, _, text) in enumerate(SUMMARY, 3):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if text == "TL;DR" or text.startswith("Consistent patterns") or text.startswith("Recommended"):
            cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)


def main() -> Path:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_summary(wb)
    write_sheet(wb, "Summit", SUMMIT)
    write_sheet(wb, "Cuyahoga", CUYAHOGA)
    write_sheet(wb, "Stark", STARK)

    out_path = OUT_DIR / f"ohio_distress_sources_{datetime.now():%Y%m%d}.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path}")
    return out_path


if __name__ == "__main__":
    main()
