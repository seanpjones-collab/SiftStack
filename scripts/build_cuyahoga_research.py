"""Build Cuyahoga County market research workbook from Market Finder JSON.

Mirrors the structure of output/ftm_county_research/Summit_County_OH_Market_Research.xlsx
exactly: 7 sheets, same column headers, same calculated fields, same star-rating
methodology. All narrative sections (Executive Summary B, Economic Indicators,
Crime, Investment Recs, Data Sources) are populated with researched values
captured in CUYAHOGA_RESEARCH below — re-runnable end-to-end.

Sources captured at retrieval (April 2026):
  - Census QuickFacts / Cuyahoga County Planning Commission population estimates
  - BLS Cleveland-Elyria MSA / FRED CLEV439* series
  - Redfin Cuyahoga County housing market page
  - Cleveland PD / Axios / Ideastream / NeighborhoodScout 2024-2025 crime data
  - FRED MEDDAYONMARUS (March 2026 reading: 55 days)

Usage:
    python scripts/build_cuyahoga_research.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO = Path(__file__).resolve().parent.parent
JSON_PATH = REPO / "output" / "market_finder_Ohio_Cuyahoga_20260426_230703.json"
OUT_PATH = REPO / "output" / "ftm_county_research" / "Cuyahoga_County_OH_Market_Research.xlsx"

# FRED MEDDAYONMARUS — most recent reading available is March 2026 = 55 days.
# (The Summit workbook used 45 from a July-2024 reading; market has slowed since.)
FRED_DOM_BASELINE = 55
FRED_DOM_BASELINE_DATE = "March 2026"

# All researched values live here so the script stays re-runnable and auditable.
CUYAHOGA_RESEARCH = {
    # ── County overview (used in Executive Summary A + Economic B + Sources) ──
    "population_2024": "1,245,873",
    "population_change_since_2020": "-1.5% since 2020 Census (1,264,817)",
    "population_density": "~2,750/sq mi",  # 1.25M / 458 sq mi land area
    "median_age": "40.9 years",
    "median_household_income": "$64,468",
    "median_household_income_note": "Census ACS 2024 (Data USA tabulation)",
    "per_capita_income": "$40,038",
    "per_capita_income_note": "Census ACS 2024 (estimate)",
    "poverty_rate": "14.7%",
    "poverty_rate_note": "Persons below poverty line, 2024 ACS — above OH ~13.4%",
    "white_non_hispanic_pct": "56.5%",
    "black_pct": "28.5%",
    "hispanic_pct": "7.0%",
    "two_or_more_pct": "4.1%",
    "asian_pct": "3.3%",
    "bachelors_or_higher": "~33%",
    "bachelors_note": "ACS 2024 — county-wide; Cleveland city only ~20%",
    "owner_occupied_pct": "59.4%",
    "owner_occupied_note": "Data USA ACS 2024 — much lower than Summit (74.1%)",
    "renter_occupied_pct": "40.6%",
    "census_median_home_value": "~$155,000",
    "census_median_home_value_note": "ACS 5-yr — lags current Sift/Redfin",

    # ── BLS Cleveland-Elyria MSA (covers Cuyahoga + Geauga + Lake + Lorain + Medina) ──
    "msa_civilian_labor_force": "~1,055,000",
    "msa_employment": "~1,011,000",
    "msa_unemployment_count": "~44,000",
    "msa_unemployment_rate": "4.2%",
    "msa_unemployment_rate_note": "BLS CLEV439URN, late 2025 — at OH state avg, below US 4.3%",
    "msa_total_nonfarm": "1,095,700",
    "msa_total_nonfarm_note": "BLS CES Dec 2025; +1,600 jobs (+0.1%) YoY",
    "msa_job_growth_forecast": "Flat to +0.5%",
    "msa_job_growth_forecast_note": "OH LMI leading indicator — healthcare-led, manufacturing softening",

    # ── Sector employment (Cleveland MSA, BLS CES Dec 2025, approximate) ──
    "sector_education_health": "~210,000",
    "sector_education_health_note": "Largest sector — Cleveland Clinic, UH, MetroHealth dominate",
    "sector_trade_transport": "~190,000",
    "sector_trade_transport_note": "Lake Erie port + I-71/I-77/I-80 logistics corridor",
    "sector_professional_business": "~155,000",
    "sector_professional_business_note": "Legal, consulting, technical services",
    "sector_government": "~120,000",
    "sector_government_note": "Federal, state, county, city, schools",
    "sector_leisure_hospitality": "~110,000",
    "sector_leisure_hospitality_note": "Restaurants, downtown sports/arts, casinos",
    "sector_manufacturing": "~110,000",
    "sector_manufacturing_note": "Steel, auto parts, plastics — slowly declining",
    "sector_financial": "~67,000",
    "sector_financial_note": "KeyBank, Progressive Insurance HQ in Mayfield Village",
    "sector_construction": "~42,000",
    "sector_construction_note": "Modest growth segment",
    "sector_other_services": "~38,000",
    "sector_information": "~14,000",

    # ── Major Employers ──
    "employers": [
        ("Cleveland Clinic", "Healthcare", "~48,000 caregivers regionally; #1 hospital US (USNWR cardiology)"),
        ("University Hospitals", "Healthcare", "~32,000 employees; 23 hospitals across NE OH"),
        ("Progressive Insurance", "Financial / Insurance", "Mayfield Village HQ; ~8,400 in Cuyahoga; largest private for-profit employer"),
        ("MetroHealth System", "Healthcare", "~8,000 employees; safety-net hospital, Brecksville campus expanding"),
        ("Cuyahoga County Government", "Government", "~7,500 employees"),
        ("Cleveland Metropolitan School District", "Education", "~7,000 employees"),
        ("KeyBank / KeyCorp", "Financial", "Cleveland HQ; ~3,500 local"),
        ("Sherwin-Williams", "Manufacturing", "Cleveland HQ; new HQ tower under construction"),
        ("Eaton Corporation", "Manufacturing", "Beachwood HQ; power management"),
        ("Case Western Reserve University", "Education", "~6,500 employees, $400M+ research"),
        ("City of Cleveland", "Government", "~7,000 employees"),
        ("Parker Hannifin", "Manufacturing", "Mayfield Heights HQ; motion/control"),
    ],

    # ── Redfin / Public housing-market data (Cuyahoga County) ──
    "redfin_median_sale_price": "$212,000",
    "redfin_median_sale_price_yoy": "+7.1% YoY (Redfin Nov 2025)",
    "redfin_median_price_per_sqft": "$140",
    "redfin_median_price_per_sqft_yoy": "Redfin 2025",
    "redfin_median_dom": "30 days",
    "redfin_median_dom_yoy": "+3 days vs 27 days prior year",
    "redfin_homes_sold_yoy": "Roughly flat YoY",
    "redfin_sale_to_list": "~98%",
    "redfin_sale_to_list_yoy": "Down ~1pp YoY",
    "redfin_above_list_pct": "~32%",
    "redfin_above_list_yoy": "Down ~3pp YoY (cooling)",
    "redfin_price_drops_pct": "~28%",
    "redfin_price_drops_yoy": "Up ~4pp YoY (more cuts)",
    "months_of_supply": "~1.4 months",  # 470 / 333 (lvg sales) approx; tight
    "months_supply_note": "Sift 470 on market vs 1,286 sold last mo — strong seller's market",

    # ── Crime data (Cleveland city = ~37% of county pop) ──
    "crime_2023_murders": 138,
    "crime_2024_murders": 105,
    "crime_2025h1_murders": "46 (pace ~92)",
    "crime_2024_homicide_change": "-24%",
    "crime_2025h1_change": "-28% vs 2024 H1 (64 → 46)",
    "crime_2024_violent": "5,663",
    "crime_2024_violent_per_100k": "1,561",
    "crime_2024_violent_change": "-13%",
    "crime_2024_felonious_assaults": "2,468",
    "crime_2024_felonious_assault_change": "Down from 3,061 peak in 2021",
    "crime_2024_robbery_change": "Down (Bibb summer plan)",
    "crime_2024_property_per_1000": "44.0",
    "crime_2024_property_per_100k": "4,401",
    "crime_2024_overall_per_1000": "59.5",
    "crime_2024_overall_per_100k": "5,953",
    "crime_2024_burglary": "3,124",
    "crime_2024_motor_vehicle_theft": "4,164",
    "crime_2024_theft": "8,794",
    "crime_2024_rape": "511",
    "crime_2024_robbery": "1,413",
    "crime_2024_aggravated_assault": "3,636",

    # ── Source URLs (for Sheet 7) ──
    "redfin_url": "redfin.com/county/2177/OH/Cuyahoga-County/housing-market",
    "redfin_county_id_note": "(Redfin internal county ID is 2177, not 418)",
}


# ── Styling ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="4472C4")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header(ws, row_idx: int, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _apply_section(ws, row_idx: int, title: str, span: int = 5) -> None:
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = SECTION_FONT
    if span > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span)


def _fmt_pct(num: float | None) -> str:
    if num is None:
        return "—"
    return f"{num:+.1%}"


def _fmt_signed_int(n: int | None) -> str:
    if n is None:
        return "—"
    return f"{n:+d}"


def _fmt_currency(v: float | None) -> str:
    if v is None or v == 0:
        return "—"
    return f"${v:,.0f}"


# ── Wholesaling score ──────────────────────────────────────────────────


def score_zip(trans: int, dom: int | None, val: float | None) -> str:
    """Reverse-engineered from Summit & Stark workbooks."""
    if not trans or dom is None:
        return "★☆☆☆☆"
    delta = dom - FRED_DOM_BASELINE
    if delta >= 10:
        return "★☆☆☆☆"
    if delta > 0:
        return "★★☆☆☆"
    if trans >= 30 and delta <= -10 and val and val < 350_000:
        return "★★★★★"
    if trans >= 20 and val and val < 400_000:
        return "★★★★☆"
    if trans >= 10:
        return "★★★☆☆"
    return "★★☆☆☆"


# ── Sheet builders ─────────────────────────────────────────────────────


def build_executive_summary(wb: Workbook, zip_rows: list[dict], nbhd_rows: list[dict]) -> None:
    R = CUYAHOGA_RESEARCH
    ws = wb.create_sheet("Executive Summary")
    ws.cell(row=1, column=1, value="CUYAHOGA COUNTY, OHIO — MARKET RESEARCH REPORT").font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws.cell(
        row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%B %d, %Y')}  |  "
              "Data Source: REI Sift Market Finder + BLS + Census + Redfin + Public Sources",
    ).font = Font(italic=True, size=10, color="595959")
    ws.merge_cells("A2:E2")

    # ── A. County Overview ────────────────────────────────────────────
    _apply_section(ws, 4, "A. COUNTY OVERVIEW")
    _apply_header(ws, 5, ["METRIC", "VALUE", "NOTES"])
    overview = [
        ("Population (2024 est.)", R["population_2024"], "2nd most populous county in Ohio (Census/Planning Commission)"),
        ("Population Change Since 2020", R["population_change_since_2020"], "Continued slow decline; outmigration"),
        ("Median Home Value (Sift)", "$162,400", "REI Sift — April 2026 (county-wide median)"),
        ("Median Sale Price (Redfin, Nov 2025)", R["redfin_median_sale_price"], R["redfin_median_sale_price_yoy"]),
        ("Median Household Income", R["median_household_income"], R["median_household_income_note"]),
        ("Per Capita Income", R["per_capita_income"], R["per_capita_income_note"]),
        ("Poverty Rate", R["poverty_rate"], R["poverty_rate_note"]),
        ("Median Age", R["median_age"], "Census 2024"),
        ("Unemployment Rate (Cleveland-Elyria MSA)", R["msa_unemployment_rate"], R["msa_unemployment_rate_note"]),
        ("Total Nonfarm Jobs (Cleveland MSA)", R["msa_total_nonfarm"], R["msa_total_nonfarm_note"]),
        ("Homes on Market", "470", "REI Sift — April 2026"),
        ("Mo. Investor Transactions", "443", "REI Sift 6-mo avg (2,659 / 6)"),
        ("Homes Sold Last Month", "1,286", "REI Sift — April 2026"),
        ("Market Rent", "$1,942/mo", "REI Sift — April 2026"),
        ("Gross Rental Yield", "10.8%", "REI Sift — April 2026 (above Summit's 9.36%)"),
        ("Homeownership Rate", "59.4%", f"Renters: 40.6% — much lower owner share than Summit's 74.1%"),
        ("Months of Supply", R["months_of_supply"], R["months_supply_note"]),
        (f"National DOM Baseline (FRED)", f"{FRED_DOM_BASELINE} days", f"FRED MEDDAYONMARUS, {FRED_DOM_BASELINE_DATE}"),
        ("Sift Median DOM (county)", "~52 days", f"3 days below national baseline"),
    ]
    for i, (m, v, n) in enumerate(overview, start=6):
        ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=n).alignment = LEFT

    # ── B. Market Assessment ──────────────────────────────────────────
    assess_row = 6 + len(overview) + 2
    _apply_section(ws, assess_row, "B. MARKET ASSESSMENT")
    _apply_header(ws, assess_row + 1, ["CATEGORY", "RATING", "COMMENTARY"])
    total_trans = sum((r.get("total_inv_trans_6mo") or 0) for r in zip_rows)
    assessments = [
        ("Investor Activity", "VERY STRONG",
         f"~443 monthly transactions ({total_trans} 6-mo) across 70 active ZIPs — 2.2x Summit's volume"),
        ("Market Velocity", "STRONG",
         "~52 day median DOM, 3 days below national baseline; Cleveland city DOM 30 days (Redfin Nov 2025)"),
        ("Price Appreciation", "POSITIVE",
         "+7.1% YoY median sale price (Redfin Nov 2025); slowing from 2024 peak but still positive"),
        ("Population", "DECLINING",
         "1.25M residents but -1.5% since 2020; outmigration to Geauga/Lorain — supply of estate sales"),
        ("Employment", "STABLE",
         "4.2% MSA unemployment (at OH avg); healthcare-led — Cleveland Clinic + UH expanding"),
        ("Crime Trend", "IMPROVING",
         "Cleveland murders -24% in 2024 (138→105), -28% H1 2025; violent crime -13%"),
        ("Rental Market", "VERY STRONG",
         "$1,942/mo median rent; 10.8% gross yield — best in OH for buy-hold + wholetail"),
        ("Price Accessibility", "VERY HIGH",
         "Median home value $162K — deeply affordable; East-side ZIPs <$100K with high investor activity"),
    ]
    for i, (cat, rating, commentary) in enumerate(assessments, start=assess_row + 2):
        ws.cell(row=i, column=1, value=cat).font = Font(bold=True)
        ws.cell(row=i, column=2, value=rating).font = Font(bold=True, color="00B050")
        ws.cell(row=i, column=3, value=commentary).alignment = LEFT

    # ── C. Top 5 ZIPs ─────────────────────────────────────────────────
    top5_row = assess_row + 2 + len(assessments) + 2
    _apply_section(ws, top5_row, "C. TOP 5 ZIP CODES FOR WHOLESALING")
    _apply_header(ws, top5_row + 1, ["RANK", "ZIP CODE", "6-MO INV TRANS", "MEDIAN HOME VALUE", "MEDIAN DOM"])
    for i, r in enumerate(zip_rows[:5], start=1):
        ws.cell(row=top5_row + 1 + i, column=1, value=i)
        ws.cell(row=top5_row + 1 + i, column=2, value=r["zip_code"])
        ws.cell(row=top5_row + 1 + i, column=3, value=r.get("total_inv_trans_6mo") or 0)
        ws.cell(row=top5_row + 1 + i, column=4, value=_fmt_currency(r.get("median_home_value")))
        ws.cell(row=top5_row + 1 + i, column=5, value=r.get("median_days_on_market") or "—")

    # ── D. Top 5 Neighborhoods ────────────────────────────────────────
    nbhd_top_row = top5_row + 1 + 5 + 2
    _apply_section(ws, nbhd_top_row, "D. TOP 5 NEIGHBORHOODS FOR WHOLESALING")
    _apply_header(ws, nbhd_top_row + 1, ["RANK", "NEIGHBORHOOD", "6-MO INV TRANS", "MEDIAN HOME VALUE", "MEDIAN DOM"])
    for i, r in enumerate(nbhd_rows[:5], start=1):
        ws.cell(row=nbhd_top_row + 1 + i, column=1, value=i)
        ws.cell(row=nbhd_top_row + 1 + i, column=2, value=r["neighborhood"])
        ws.cell(row=nbhd_top_row + 1 + i, column=3, value=r.get("total_inv_trans_6mo") or 0)
        ws.cell(row=nbhd_top_row + 1 + i, column=4, value=_fmt_currency(r.get("median_home_value")))
        ws.cell(row=nbhd_top_row + 1 + i, column=5, value=r.get("median_days_on_market") or "—")


def build_zip_analysis(wb: Workbook, zip_rows: list[dict]) -> None:
    ws = wb.create_sheet("ZIP Code Analysis")
    ws.cell(row=1, column=1, value=f"CUYAHOGA COUNTY, OH — ZIP CODE ANALYSIS ({len(zip_rows)} ZIP Codes)").font = TITLE_FONT
    ws.merge_cells("A1:K1")
    headers = [
        "ZIP CODE", "6-MO INV TRANS", "HOMES ON MARKET", "HOMES SOLD/MO",
        "MEDIAN DOM", "DOM vs NATIONAL", "MEDIAN HOME VALUE", "MEDIAN SALE PRICE",
        "SPREAD %", "SUPPLY MONTHS", "WHOLESALING SCORE",
    ]
    _apply_header(ws, 2, headers)
    for i, r in enumerate(zip_rows, start=3):
        trans = r.get("total_inv_trans_6mo") or 0
        on_mkt = r.get("homes_on_market") or 0
        sold = r.get("homes_sold_last_month") or 0
        dom = r.get("median_days_on_market")
        val = r.get("median_home_value")
        sale = r.get("median_sale_price")
        dom_delta = (dom - FRED_DOM_BASELINE) if dom else None
        spread = ((sale - val) / val) if (sale and val) else None
        supply = (on_mkt / sold) if sold else None
        score = score_zip(trans, dom, val)
        row = [
            r["zip_code"], trans, on_mkt, sold,
            dom if dom is not None else "—",
            _fmt_signed_int(dom_delta) if dom_delta is not None else "—",
            _fmt_currency(val), _fmt_currency(sale),
            _fmt_pct(spread) if spread is not None else "—",
            round(supply, 1) if supply is not None else "—",
            score,
        ]
        for col_idx, val_ in enumerate(row, start=1):
            ws.cell(row=i, column=col_idx, value=val_)
    ws.freeze_panes = "A3"


def build_neighborhood_analysis(wb: Workbook, nbhd_rows: list[dict]) -> None:
    ws = wb.create_sheet("Neighborhood Analysis")
    ws.cell(row=1, column=1, value=f"CUYAHOGA COUNTY, OH — NEIGHBORHOOD ANALYSIS ({len(nbhd_rows)} Neighborhoods)").font = TITLE_FONT
    ws.merge_cells("A1:K1")
    headers_n = [
        "NEIGHBORHOOD", "6-MO INV TRANS", "HOMES ON MARKET", "HOMES SOLD/MO",
        "MEDIAN DOM", "DOM vs NATIONAL", "MEDIAN HOME VALUE", "MEDIAN SALE PRICE",
        "SPREAD %", "SUPPLY MONTHS", "WHOLESALING SCORE",
    ]
    _apply_header(ws, 2, headers_n)
    for i, r in enumerate(nbhd_rows, start=3):
        trans = r.get("total_inv_trans_6mo") or 0
        on_mkt = r.get("homes_on_market") or 0
        sold = r.get("homes_sold_last_month") or 0
        dom = r.get("median_days_on_market")
        val = r.get("median_home_value")
        sale = r.get("median_sale_price")
        dom_delta = (dom - FRED_DOM_BASELINE) if dom else None
        spread = ((sale - val) / val) if (sale and val) else None
        supply = (on_mkt / sold) if sold else None
        score = score_zip(trans, dom, val)
        row = [
            r["neighborhood"], trans, on_mkt, sold,
            dom if dom is not None else "—",
            _fmt_signed_int(dom_delta) if dom_delta is not None else "—",
            _fmt_currency(val), _fmt_currency(sale),
            _fmt_pct(spread) if spread is not None else "—",
            round(supply, 1) if supply is not None else "—",
            score,
        ]
        for col_idx, val_ in enumerate(row, start=1):
            ws.cell(row=i, column=col_idx, value=val_)
    ws.freeze_panes = "A3"


def build_economic_indicators(wb: Workbook) -> None:
    R = CUYAHOGA_RESEARCH
    ws = wb.create_sheet("Economic Indicators")
    ws.cell(row=1, column=1, value="CUYAHOGA COUNTY / CLEVELAND-ELYRIA MSA — ECONOMIC INDICATORS").font = TITLE_FONT
    ws.merge_cells("A1:C1")

    # A. Employment data
    _apply_section(ws, 3, "A. EMPLOYMENT DATA (Cleveland-Elyria MSA — BLS)", span=3)
    _apply_header(ws, 4, ["METRIC", "VALUE", "NOTES"])
    employment = [
        ("Civilian Labor Force", R["msa_civilian_labor_force"], "Cleveland-Elyria MSA (Cuyahoga + 4 surrounding counties)"),
        ("Employment", R["msa_employment"], "Cleveland-Elyria MSA"),
        ("Unemployment", R["msa_unemployment_count"], "Cleveland-Elyria MSA"),
        ("Unemployment Rate", R["msa_unemployment_rate"], R["msa_unemployment_rate_note"]),
        ("Total Nonfarm Payroll Jobs", R["msa_total_nonfarm"], R["msa_total_nonfarm_note"]),
        ("6-Mo Job Growth Forecast", R["msa_job_growth_forecast"], R["msa_job_growth_forecast_note"]),
    ]
    for i, (m, v, n) in enumerate(employment, start=5):
        ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=n).alignment = LEFT

    # Sector breakdown
    sector_row = 5 + len(employment) + 2
    ws.cell(row=sector_row, column=1, value="EMPLOYMENT BY SECTOR (Cleveland-Elyria MSA)").font = Font(bold=True, size=11)
    _apply_header(ws, sector_row + 1, ["SECTOR", "APPROX JOBS", "NOTES"])
    sectors = [
        ("Education & Health Services", R["sector_education_health"], R["sector_education_health_note"]),
        ("Trade, Transport & Utilities", R["sector_trade_transport"], R["sector_trade_transport_note"]),
        ("Professional & Business Svcs", R["sector_professional_business"], R["sector_professional_business_note"]),
        ("Government", R["sector_government"], R["sector_government_note"]),
        ("Leisure & Hospitality", R["sector_leisure_hospitality"], R["sector_leisure_hospitality_note"]),
        ("Manufacturing", R["sector_manufacturing"], R["sector_manufacturing_note"]),
        ("Financial Activities", R["sector_financial"], R["sector_financial_note"]),
        ("Construction", R["sector_construction"], R["sector_construction_note"]),
        ("Other Services", R["sector_other_services"], "Repair, personal services, civic orgs"),
        ("Information", R["sector_information"], "Media, telecom, tech"),
    ]
    for i, (s, j, n) in enumerate(sectors, start=sector_row + 2):
        ws.cell(row=i, column=1, value=s).font = Font(bold=True)
        ws.cell(row=i, column=2, value=j)
        ws.cell(row=i, column=3, value=n).alignment = LEFT

    # Major employers
    emp_row = sector_row + 2 + len(sectors) + 2
    ws.cell(row=emp_row, column=1, value="MAJOR EMPLOYERS").font = Font(bold=True, size=11)
    _apply_header(ws, emp_row + 1, ["EMPLOYER", "SECTOR", "NOTES"])
    for i, (name, sector, note) in enumerate(R["employers"], start=emp_row + 2):
        ws.cell(row=i, column=1, value=name).font = Font(bold=True)
        ws.cell(row=i, column=2, value=sector)
        ws.cell(row=i, column=3, value=note).alignment = LEFT

    # B. Demographic data
    demo_row = emp_row + 2 + len(R["employers"]) + 2
    _apply_section(ws, demo_row, "B. DEMOGRAPHIC DATA (Census 2024 / ACS)", span=3)
    _apply_header(ws, demo_row + 1, ["METRIC", "VALUE", "NOTES"])
    demographics = [
        ("Population (2024)", R["population_2024"], "2nd most populous county in Ohio"),
        ("Population Change Since 2020", R["population_change_since_2020"], "Continued out-migration"),
        ("Population Density", R["population_density"], "Densely urban — Cleveland metro core"),
        ("Median Age", R["median_age"], "Census 2024"),
        ("Median Household Income", R["median_household_income"], R["median_household_income_note"]),
        ("Per Capita Income", R["per_capita_income"], R["per_capita_income_note"]),
        ("Poverty Rate", R["poverty_rate"], R["poverty_rate_note"]),
        ("White (non-Hispanic)", R["white_non_hispanic_pct"], "Largest group; lower share than Summit (73.4%)"),
        ("Black / African American", R["black_pct"], "Cleveland city is ~50% Black; concentrated East Side"),
        ("Hispanic / Latino", R["hispanic_pct"], "Heavily concentrated in 44102 / Stockyards / Clark-Fulton"),
        ("Two or more races", R["two_or_more_pct"], "Growing multiracial population"),
        ("Asian", R["asian_pct"], "Concentrated University Circle / Solon"),
        ("Bachelor's Degree or Higher", R["bachelors_or_higher"], R["bachelors_note"]),
        ("Owner-Occupied Housing Units", R["owner_occupied_pct"], R["owner_occupied_note"]),
        ("Renter-Occupied Housing Units", R["renter_occupied_pct"], "High rental demand, big BTR opportunity"),
        ("Median Home Value (Census ACS)", R["census_median_home_value"], R["census_median_home_value_note"]),
    ]
    for i, (m, v, n) in enumerate(demographics, start=demo_row + 2):
        ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=n).alignment = LEFT

    # C. Housing market data
    hm_row = demo_row + 2 + len(demographics) + 2
    _apply_section(ws, hm_row, "C. HOUSING MARKET DATA (Redfin / Public Sources)", span=3)
    _apply_header(ws, hm_row + 1, ["METRIC", "VALUE", "YoY CHANGE"])
    housing = [
        ("Median Sale Price (Cuyahoga Co.)", R["redfin_median_sale_price"], R["redfin_median_sale_price_yoy"]),
        ("Median Price per Sq Ft", R["redfin_median_price_per_sqft"], R["redfin_median_price_per_sqft_yoy"]),
        ("Median Days on Market (county)", R["redfin_median_dom"], R["redfin_median_dom_yoy"]),
        ("Homes Sold (Sift, last month)", "1,286", "Strong sales velocity"),
        ("Months of Supply", R["months_of_supply"], R["months_supply_note"]),
        ("Sale-to-List Price Ratio", R["redfin_sale_to_list"], R["redfin_sale_to_list_yoy"]),
        ("Homes Sold Above List Price", R["redfin_above_list_pct"], R["redfin_above_list_yoy"]),
        ("Homes with Price Drops", R["redfin_price_drops_pct"], R["redfin_price_drops_yoy"]),
        ("Forecasted Price Growth", "+3-5%", "Through end of 2026 (consensus, slowing)"),
        ("Property Type — Single Family", "87.0%", "REI Sift summary panel — dominant"),
        ("Property Type — Multi-Family", "~5%", "Limited; Cleveland has 2-4 unit dense pockets"),
        ("Year Built — 2000+", "6.7%", "Almost no new construction; high distress pool"),
        ("Year Built — Pre-1939", "~35%", "Heavy estate/probate inventory; lead/asbestos risk"),
        ("Year Built — 1940-1969", "~35%", "Largest cohort — postwar bungalows + Colonials"),
        ("Year Built — 1970-1999", "~23%", "Suburban ranch + tract"),
    ]
    for i, (m, v, c) in enumerate(housing, start=hm_row + 2):
        ws.cell(row=i, column=1, value=m).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=c).alignment = LEFT


def build_crime_safety(wb: Workbook) -> None:
    R = CUYAHOGA_RESEARCH
    ws = wb.create_sheet("Crime & Safety")
    ws.cell(row=1, column=1, value="CUYAHOGA COUNTY / CLEVELAND, OH — CRIME & SAFETY ANALYSIS").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # A. Crime stats
    _apply_section(ws, 3, "A. CRIME STATISTICS — CLEVELAND CITY (Most Recent FBI/CPD Data)", span=4)
    _apply_header(ws, 4, ["CRIME TYPE", "2023 COUNT", "2024 COUNT", "CHANGE"])
    crime_rows = [
        ("Murders", "138", str(R["crime_2024_murders"]), R["crime_2024_homicide_change"]),
        ("Non-Fatal Shootings (Felonious Assault)", "~3,000", R["crime_2024_felonious_assaults"], "Down sharply from 2021 peak"),
        ("Robberies", "~1,600", R["crime_2024_robbery"], "Down ~12%"),
        ("Aggravated Assaults", "~3,900", R["crime_2024_aggravated_assault"], "Down ~7%"),
        ("Rapes (Reported)", "~545", R["crime_2024_rape"], "Down ~6%"),
        ("Motor Vehicle Thefts", "~5,200", R["crime_2024_motor_vehicle_theft"], "Down ~20% (Hyundai/Kia thefts cooling)"),
        ("Burglaries", "~3,400", R["crime_2024_burglary"], "Down ~8%"),
        ("Larceny / Theft", "~9,500", R["crime_2024_theft"], "Down ~7%"),
        ("Total Violent Crimes", "~6,500", R["crime_2024_violent"], R["crime_2024_violent_change"]),
        ("Total Property Crimes", "~17,500", "16,082", "~-8%"),
        ("Overall Crime Rate/100K", "~6,500", R["crime_2024_overall_per_100k"], "~-8%"),
    ]
    for i, (typ, p, c, ch) in enumerate(crime_rows, start=5):
        ws.cell(row=i, column=1, value=typ).font = Font(bold=True)
        ws.cell(row=i, column=2, value=p)
        ws.cell(row=i, column=3, value=c)
        ws.cell(row=i, column=4, value=ch)

    # B. Historical murder trend
    hist_row = 5 + len(crime_rows) + 2
    _apply_section(ws, hist_row, "B. HISTORICAL MURDER TREND — CLEVELAND CITY", span=4)
    _apply_header(ws, hist_row + 1, ["YEAR", "MURDERS", "YoY CHANGE", "NOTES"])
    hist = [
        ("2020", "180", "—", "Pandemic-era peak (5-year high)"),
        ("2021", "165", "-8%", "Still elevated"),
        ("2022", "150", "-9%", "Gradual decline"),
        ("2023", "138", "-8%", "Continued improvement"),
        ("2024", str(R["crime_2024_murders"]), R["crime_2024_homicide_change"], "Bibb summer-safety plan; lowest in 5 years"),
        ("2025 H1", R["crime_2025h1_murders"], R["crime_2025h1_change"], "On track for lowest year of decade"),
    ]
    for i, (y, m, ch, n) in enumerate(hist, start=hist_row + 2):
        ws.cell(row=i, column=1, value=y).font = Font(bold=True)
        ws.cell(row=i, column=2, value=m)
        ws.cell(row=i, column=3, value=ch)
        ws.cell(row=i, column=4, value=n).alignment = LEFT

    # C. Safety assessment by area
    safety_row = hist_row + 2 + len(hist) + 2
    _apply_section(ws, safety_row, "C. SAFETY ASSESSMENT BY AREA", span=4)
    _apply_header(ws, safety_row + 1, ["AREA", "SAFETY RATING", "DOM CONTEXT", "INVESTOR NOTES"])
    areas = [
        ("East Side Cleveland (44104, 44105, 44108, 44120, 44128)", "Lower",
         "55-64 days",
         "Highest investor activity county-wide — 44105 leads at 225 trans. Deep discounts; experienced buyers required."),
        ("West Side Cleveland (44102, 44109, 44111, 44135)", "Moderate",
         "41-54 days",
         "Strong wholesaling pocket — 44102 (217 trans), 44111 (86 trans). Faster DOM than East Side."),
        ("Downtown / University Circle (44103, 44106, 44114, 44115)", "Moderate-Lower",
         "55-75 days",
         "Mixed — University Circle premium pocket; Downtown thin investor activity. Niche flips, not wholesale."),
        ("Inner Suburbs — Lakewood, Cleveland Hts, S. Euclid (44107, 44118, 44121)", "Moderate-High",
         "44-56 days",
         "Lakewood (44107) — fast DOM, mid prices. Cleveland Hts (44118) — large stock, motivated estate sales."),
        ("Inner Suburbs — Parma, Garfield Hts, Maple Hts (44129, 44134, 44125, 44137)", "Moderate-High",
         "38-55 days",
         "Postwar bungalow belt — 44125, 44137, 44129 all 4-star. Best risk/reward in county."),
        ("Outer Suburbs — Strongsville, North Royalton (44136, 44133, 44147)", "High",
         "45-55 days",
         "Strongsville violent crime 82% below national. Higher prices ($245K+); flips/wholetail not pure wholesale."),
        ("Outer Suburbs — Westlake, Bay Village, Rocky River (44116, 44140, 44145)", "Very High",
         "47-72 days",
         "Westside premium — $400K+ medians; very limited wholesale spread; cash-buyer hold list."),
        ("Outer Suburbs — Solon, Pepper Pike, Hunting Valley (44022, 44124, 44139)", "Very High",
         "53-90 days",
         "Affluent East suburbs — Solon $350K median. Estate-sale pocket but minimal margins; relationship plays."),
        ("Brook Park / Berea / Olmsted Falls (44017, 44070, 44138, 44142)", "High",
         "38-43 days",
         "All 4-star — Berea (44017), N. Olmsted (44070), Brook Park (44142). Solid suburban activity."),
        ("East Cleveland (city, 44112)", "Very Low",
         "69 days",
         "Distinct from Cleveland Hts — extreme distress, but DOM 14 above baseline. Very experienced only."),
    ]
    for i, (area, rating, dom, note) in enumerate(areas, start=safety_row + 2):
        ws.cell(row=i, column=1, value=area).font = Font(bold=True)
        ws.cell(row=i, column=2, value=rating)
        ws.cell(row=i, column=3, value=dom)
        ws.cell(row=i, column=4, value=note).alignment = LEFT

    # D. Key insights
    insights_row = safety_row + 2 + len(areas) + 2
    _apply_section(ws, insights_row, "D. KEY INSIGHTS FOR INVESTORS", span=4)
    insights = [
        "• Cleveland violent crime is in a clear DOWN trend — murders -24% in 2024, -28% H1 2025. Best fundamentals in 5 years; buyer confidence improving.",
        "• East Side ZIPs (44105, 44128, 44108, 44104) carry the deepest discounts AND highest investor velocity — these are wholesale-only plays. Expect 1 in 17 victimization rate; buyers must be experienced operators with rental portfolios.",
        "• West Side (44102, 44111, 44109) is the optimal balance of activity + safety — DOM 41-54 days, mid-tier prices. Most accessible to first-time wholesale buyers.",
        "• Inner-ring postwar belt (44125, 44129, 44134, 44137, 44146) offers the best risk-adjusted returns — moderate crime, fast DOM (~38-55), $130-210K medians. This is your TIER 1 target.",
        "• Cleveland's overall crime rate (5,953/100K) is 78% above national average — this is the structural reason home values are so deeply discounted vs the rest of the country and why wholesale spreads are wide.",
        "• High-end East suburbs (Solon, Hunting Valley, Pepper Pike) and West suburbs (Westlake, Bay Village) are too expensive for wholesale margin but should be tracked for cash-buyer relationships and high-end estate sales.",
        "• East Cleveland (44112) is a distinct municipality from Cleveland Heights — extreme distress with DOM 14 days above national baseline. Only experienced operators with hold capacity should target.",
    ]
    for i, txt in enumerate(insights, start=insights_row + 1):
        ws.cell(row=i, column=1, value=txt).alignment = LEFT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)


def build_investment_recommendations(wb: Workbook) -> None:
    """Tier 1/2 ZIPs and neighborhoods are curated from the ZIP/Neighborhood
    Analysis sheets — the rationale text is hand-written, so we don't pass the
    raw row data through here. Update tier1_zips / tier1_nbhds / tier2 below
    if the underlying Sift data shifts materially."""
    ws = wb.create_sheet("Investment Recommendations")
    ws.cell(row=1, column=1, value="CUYAHOGA COUNTY, OH — INVESTMENT RECOMMENDATIONS").font = TITLE_FONT
    ws.merge_cells("A1:E1")

    # A. Tier 1 — Highest priority ZIPs (curated from 4-star list, top by activity volume)
    _apply_section(ws, 3, "A. TIER 1 — HIGHEST PRIORITY ZIP CODES")
    _apply_header(ws, 4, ["ZIP CODE", "INV TRANS (6mo)", "MEDIAN VALUE", "MEDIAN DOM", "RATIONALE"])
    tier1_zips = [
        ("44111", 86, "$167,776", "41 days",
         "West Park / Riverside — DOM 14 below baseline, mid-tier accessible price, 86 trans. TOP PICK for new wholesale buyers."),
        ("44110", 65, "$102,005", "43 days",
         "Collinwood — ultra-affordable ($102K), DOM 12 below baseline, deep distress pool. East Side wholesale leader."),
        ("44129", 56, "$210,087", "38 days",
         "Parma SW — DOM 17 below baseline (fastest sub-50 trans group), strong middle-class buyer pool. EXCELLENT."),
        ("44132", 46, "$151,331", "45 days",
         "Euclid — affordable ($151K), DOM 10 below baseline, 46 trans. Solid Tier 1 with active buyer pool."),
        ("44134", 40, "$208,055", "45 days",
         "Parma — 40 trans, mid-price ($208K), DOM 10 below baseline. Inner-ring postwar bungalow heaven."),
        ("44107", 38, "$289,111", "44 days",
         "Lakewood — high investor demand, fast DOM, denser walkable; trends wholetail rather than pure wholesale at $289K."),
        ("44138", 35, "$268,712", "43 days",
         "Olmsted Falls / Olmsted Township — fast DOM, mid-price. Suburban quality with consistent activity."),
        ("44070", 34, "$267,252", "43 days",
         "North Olmsted — DOM 12 below baseline, 34 trans, $267K. Solid suburban Tier 1."),
    ]
    for i, (z, t, v, d, r) in enumerate(tier1_zips, start=5):
        ws.cell(row=i, column=1, value=z).font = Font(bold=True)
        ws.cell(row=i, column=2, value=t)
        ws.cell(row=i, column=3, value=v)
        ws.cell(row=i, column=4, value=d)
        ws.cell(row=i, column=5, value=r).alignment = LEFT

    # B. Tier 1 neighborhoods
    nbhd_row = 5 + len(tier1_zips) + 2
    _apply_section(ws, nbhd_row, "B. TIER 1 — HIGHEST PRIORITY NEIGHBORHOODS")
    _apply_header(ws, nbhd_row + 1, ["NEIGHBORHOOD", "INV TRANS", "MEDIAN VALUE", "MEDIAN DOM", "RATIONALE"])
    tier1_nbhds = [
        ("Edgewater - Cleveland", 125, "$247,179", "49 days",
         "#1 neighborhood county-wide — 125 trans (4x next), DOM 6 below baseline. Lakewood-adjacent gentrifying core."),
        ("Maple Heights East", 29, "$134,975", "45 days",
         "29 trans, sub-$135K, DOM 10 below baseline. Inner-ring postwar bungalows. Excellent risk/reward."),
        ("Lee-Harvard", 26, "$133,890", "57 days",
         "26 trans, very affordable ($134K). DOM slightly above baseline → motivated sellers; deep estate inventory."),
        ("Garfield Heights Northeast", 26, "$110,588", "63 days",
         "Ultra-affordable ($110K) with high DOM (63 days = motivated sellers). High distress, strong wholesale spread."),
        ("Union-Miles", 25, "$92,584", "48 days",
         "Sub-$100K median + DOM below baseline = pure wholesale territory. Experienced operators / cash buyers only."),
        ("City Center - Maple Heights", 26, "$136,062", "66 days",
         "26 trans, DOM 11 above baseline = motivated sellers. Affordable Maple Heights core."),
    ]
    for i, (n, t, v, d, r) in enumerate(tier1_nbhds, start=nbhd_row + 2):
        ws.cell(row=i, column=1, value=n).font = Font(bold=True)
        ws.cell(row=i, column=2, value=t)
        ws.cell(row=i, column=3, value=v)
        ws.cell(row=i, column=4, value=d)
        ws.cell(row=i, column=5, value=r).alignment = LEFT

    # C. Tier 2 — Secondary opportunities
    tier2_row = nbhd_row + 2 + len(tier1_nbhds) + 2
    _apply_section(ws, tier2_row, "C. TIER 2 — SECONDARY OPPORTUNITIES")
    _apply_header(ws, tier2_row + 1, ["ZIP CODE", "INV TRANS", "MEDIAN VALUE", "MEDIAN DOM", "RATIONALE"])
    tier2 = [
        ("44105", 225, "$96,486", "61 days",
         "Highest county volume (225 trans) BUT DOM 6 above baseline + ultra-low price. Slept Slavic Village — very experienced operators only; high crime."),
        ("44102", 217, "$138,122", "54 days",
         "2nd highest volume, DOM 1 below baseline. Stockyards / Clark-Fulton — Hispanic neighborhood, multi-family heavy. Solid wholesale."),
        ("44128", 183, "$116,936", "55 days",
         "183 trans, sub-$120K, DOM at baseline. East Side Lee-Miles — high distress, wide spread."),
        ("44120", 125, "$169,494", "61 days",
         "Shaker Heights overlap (Shaker Hts side is safer) — 125 trans, DOM slightly slow → motivated sellers."),
        ("44125", 123, "$146,024", "53 days",
         "Garfield Heights / Cuyahoga Heights — affordable, DOM near baseline. Solid Tier 2 with industrial-adjacent housing."),
        ("44137", 109, "$137,931", "48 days",
         "Maple Heights — strong activity, fast DOM, affordable. Inner-ring with active buyer pool."),
    ]
    for i, (z, t, v, d, r) in enumerate(tier2, start=tier2_row + 2):
        ws.cell(row=i, column=1, value=z).font = Font(bold=True)
        ws.cell(row=i, column=2, value=t)
        ws.cell(row=i, column=3, value=v)
        ws.cell(row=i, column=4, value=d)
        ws.cell(row=i, column=5, value=r).alignment = LEFT

    # D. Market timing
    timing_row = tier2_row + 2 + len(tier2) + 2
    _apply_section(ws, timing_row, "D. MARKET TIMING CONSIDERATIONS")
    _apply_header(ws, timing_row + 1, ["FACTOR", "", "CURRENT STATUS", "IMPLICATION"])
    timing = [
        ("Price Trend", "", "+7.1% YoY (Redfin Nov 2025)",
         "Slowing from 2024 highs but still positive — move quickly on accepted offers; ARV inflation favors wholesalers"),
        ("Inventory", "", "470 homes on market (Sift)",
         "Tight — 1.4-month supply; sellers have leverage on retail but motivated sellers still abundant via court records"),
        ("Days on Market", "", "Sift ~52 / Redfin 30 days",
         "Faster than national 55-day baseline; market velocity is healthy for assignment exits"),
        ("Distress Level", "", "Pre-1970 stock dominant; -1.5% pop since 2020",
         "Massive distress pool — estates, probates, deferred maintenance, lead/asbestos. First-to-market wins."),
        ("Employment", "", "4.2% MSA unemployment",
         "Stable — at OH avg, below US 4.3%. Healthcare-led, recession-resistant; manufacturing softening"),
        ("Rental Yield", "", "10.8% gross yield ($1,942/mo rent)",
         "BEST yield in OH; Cuyahoga is the #1 BTR / buy-hold market in state — strong cash buyer demand"),
        ("Crime Trend", "", "Murders -24% YoY; -28% H1 2025",
         "Improving fundamentals; positive for ARV growth and out-of-state buyer confidence"),
    ]
    for i, (factor, _, status, impl) in enumerate(timing, start=timing_row + 2):
        ws.cell(row=i, column=1, value=factor).font = Font(bold=True)
        ws.cell(row=i, column=3, value=status).alignment = LEFT
        ws.cell(row=i, column=4, value=impl).alignment = LEFT

    # E. Recommended strategy
    strat_row = timing_row + 2 + len(timing) + 2
    _apply_section(ws, strat_row, "E. RECOMMENDED STRATEGY")
    strategy = [
        "1. FOCUS DIRECT MAIL on Tier 1 ZIPs 44111, 44110, 44129, 44132, 44134, 44107, 44138, 44070 — these 8 ZIPs combine fast DOM (38-45 days), accessible prices ($102K-$289K), and strong investor velocity (20-86 trans each).",
        "2. PRIMARY NEIGHBORHOODS: Edgewater, Maple Heights East, Lee-Harvard, Garfield Heights NE, Union-Miles, City Center-Maple Heights — these 6 capture the best inner-ring postwar bungalow inventory.",
        "3. PRICE RANGE: Acquisition target $80K-$220K — captures the bulk of motivated-seller inventory, ensures strong cash-buyer demand for assignments, and aligns with Cuyahoga's median $162K Sift value.",
        "4. DOM TRIGGER: Properties sitting 60+ days on market are prime outreach targets — motivated sellers exist in 44105 (61 days), 44108 (64), 44112 (69), 44120 (61), Lee-Harvard (57), and Garfield Hts NE (63).",
        "5. DISTRESS LISTS: Cuyahoga is a SiftStack primary county — DAILY foreclosure pulls from cpdocket + DLN + probate court (62% address-fill rate via MyPlace lookup) feed direct mail. This is your unfair advantage vs MLS-only competitors.",
        "6. EXIT STRATEGY: 10.8% gross rental yield + $1,942/mo rents = strongest BTR / cash-buyer demand in OH. Wholetail and subject-to are equally viable alongside straight assignments — diversify exit options per deal.",
        "7. AVOID for wholesale: 44022 ($619K), 44023 ($1.16M), 44040 ($702K), 44141 ($424K), 44145 ($416K), 44139 ($438K) — values too high for wholesale spread. Track these for cash-buyer relationships and high-end estate referrals only.",
        "8. EAST CLEVELAND CAVEAT: 44112 is a separate distressed municipality (NOT Cleveland Heights) — DOM 14 days above baseline, ultra-low values. High wholesale spread but only experienced operators with hold capital should engage.",
    ]
    for i, txt in enumerate(strategy, start=strat_row + 1):
        ws.cell(row=i, column=1, value=txt).alignment = LEFT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)


def build_data_sources(wb: Workbook) -> None:
    today = datetime.now().strftime("%B %Y")
    ws = wb.create_sheet("Data Sources")
    ws.cell(row=1, column=1, value="DATA SOURCES & METHODOLOGY").font = TITLE_FONT
    ws.merge_cells("A1:D1")

    _apply_section(ws, 3, "A. PRIMARY DATA SOURCES", span=4)
    _apply_header(ws, 4, ["SOURCE", "DATA TYPE", "DATE RETRIEVED", "URL / NOTES"])
    sources = [
        ("REI Sift Market Finder", "Investor transactions, home values, DOM, rent, yield",
         "April 2026", "app.reisift.io/market-finder (extracted 2026-04-26)"),
        ("FRED (St. Louis Fed)", f"National Days on Market baseline ({FRED_DOM_BASELINE} days, {FRED_DOM_BASELINE_DATE})",
         today, "fred.stlouisfed.org/series/MEDDAYONMARUS"),
        ("Bureau of Labor Statistics", "Employment, unemployment, sector data — Cleveland-Elyria MSA",
         "Dec 2025 / April 2026", "bls.gov/eag/eag.oh_cleveland_msa.htm (direct WebFetch blocked, accessed via WebSearch)"),
        ("U.S. Census Bureau / Data USA", "Demographics, population, income, housing",
         "2024 ACS", "census.gov/quickfacts/cuyahogacountyohio (direct WebFetch blocked, accessed via Data USA + Census Reporter mirrors)"),
        ("Cuyahoga County Planning Commission", "Annual population estimates",
         "2024", "countyplanning.us/resources/census-data/population-estimates/"),
        ("Redfin", "Median sale price, price/sqft, DOM, sales volume",
         "Nov 2025", f"{CUYAHOGA_RESEARCH['redfin_url']} (direct WebFetch blocked, accessed via WebSearch summary)"),
        ("Cleveland Police / City of Cleveland", "Crime statistics, annual report",
         "2024-2025", "clevelandohio.gov/news (Bibb administration crime reports)"),
        ("NeighborhoodScout", "Crime rates per 100K population",
         "2024 (released Oct 2025)", "neighborhoodscout.com/oh/cleveland/crime"),
        ("Axios Cleveland / Ideastream", "2024 + 2025 H1 homicide trend reporting",
         "2025", "axios.com/local/cleveland; ideastream.org/law-justice"),
        ("CrimeGrade / DoorProfit", "Neighborhood-level crime grades and ZIP heatmaps",
         "2024-2025", "crimegrade.org/safest-places-in-cleveland-oh/"),
        ("Niche / Cleveland Magazine", "Best/safest suburbs rankings",
         "2026 ranking cycle", "niche.com/places-to-live/search/safest-suburbs/m/cleveland-metro-area/"),
        ("Crain's Cleveland Business", "Largest-employer rankings (Cuyahoga County)",
         "Latest available", "crainscleveland.com (rankings updated annually)"),
        ("OhioLMI.com", "Leading employment indicators",
         "Aug 2025", "ohiolmi.com"),
    ]
    for i, (s, t, d, u) in enumerate(sources, start=5):
        ws.cell(row=i, column=1, value=s).font = Font(bold=True)
        ws.cell(row=i, column=2, value=t).alignment = LEFT
        ws.cell(row=i, column=3, value=d)
        ws.cell(row=i, column=4, value=u).alignment = LEFT

    method_row = 5 + len(sources) + 2
    _apply_section(ws, method_row, "B. METHODOLOGY — CALCULATED FIELDS", span=4)
    _apply_header(ws, method_row + 1, ["FIELD", "FORMULA", "INTERPRETATION"])
    methods = [
        ("Wholesaling Score", "Composite: Inv Trans + DOM vs Nat'l + Price",
         "★★★★★ = best wholesale opportunity; ★☆☆☆☆ = avoid"),
        ("DOM vs National", f"Median DOM − FRED National Baseline ({FRED_DOM_BASELINE} days)",
         "Negative = faster than national average (green)"),
        ("Spread %", "(Median Sale Price − Median Home Value) / Med. Value",
         "Negative = buyers below list; positive = above list"),
        ("Supply Months", "Homes on Market ÷ Homes Sold Last Month",
         "<3 = seller's market; 3-6 = balanced; >6 = buyer's market"),
        ("Tier Classification", "Based on 6-mo investor transaction volume + DOM + price",
         "Tier 1 = 4-5 star ZIPs by score; Tier 2 = high-volume but caveats; Tier 3 = avoid for wholesale"),
    ]
    for i, (f, fo, n) in enumerate(methods, start=method_row + 2):
        ws.cell(row=i, column=1, value=f).font = Font(bold=True)
        ws.cell(row=i, column=2, value=fo).alignment = LEFT
        ws.cell(row=i, column=3, value=n).alignment = LEFT

    disc_row = method_row + 2 + len(methods) + 2
    _apply_section(ws, disc_row, "C. DISCLAIMERS", span=4)
    disclaimers = [
        "• Data is current as of retrieval date and subject to change; always verify before making investment decisions.",
        "• REI Sift data represents proprietary investor transaction tracking and may differ from MLS-reported figures.",
        "• Direct WebFetch to BLS, Census, Redfin, and FRED returned 403 from this environment — values were sourced via WebSearch result summaries citing those same primary sources. Cross-verify against the URLs listed above before publishing.",
        "• Crime statistics are 2024 calendar year (released Oct 2025) plus 2025 H1 partial-year reporting; subject to final FBI UCR audit.",
        "• Cleveland-Elyria MSA covers 5 counties (Cuyahoga, Geauga, Lake, Lorain, Medina). Cuyahoga is ~60% of the MSA; sector data is MSA-wide.",
        "• Cleveland city crime data (Section A) covers ONLY the city of Cleveland (~36% of county pop). Suburbs are dramatically safer — see Section C area assessment.",
        "• Investment recommendations are for informational purposes only and do not constitute financial or legal advice.",
        "• Always conduct independent due diligence, title search, and property inspection before investing.",
        f"• National DOM baseline ({FRED_DOM_BASELINE} days, {FRED_DOM_BASELINE_DATE} reading) reflects post-2024 market slowdown — verify current FRED MEDDAYONMARUS before re-running this report.",
    ]
    for i, txt in enumerate(disclaimers, start=disc_row + 1):
        ws.cell(row=i, column=1, value=txt).alignment = LEFT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)


# ── Main build ─────────────────────────────────────────────────────────


def main() -> None:
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    zip_rows = data["zip_data"]
    nbhd_rows = data["neighborhood_data"]

    # Sort by 6-mo investor transactions desc
    zip_rows.sort(key=lambda r: r.get("total_inv_trans_6mo") or 0, reverse=True)
    nbhd_rows.sort(key=lambda r: r.get("total_inv_trans_6mo") or 0, reverse=True)

    wb = Workbook()
    wb.remove(wb.active)

    build_executive_summary(wb, zip_rows, nbhd_rows)
    build_zip_analysis(wb, zip_rows)
    build_neighborhood_analysis(wb, nbhd_rows)
    build_economic_indicators(wb)
    build_crime_safety(wb)
    build_investment_recommendations(wb)
    build_data_sources(wb)

    # Column widths — wide for narrative columns, normal for numeric
    column_widths = {
        "Executive Summary": {"A": 40, "B": 22, "C": 60, "D": 22, "E": 22},
        "ZIP Code Analysis": {"A": 12, "B": 14, "C": 14, "D": 14, "E": 12, "F": 16, "G": 18, "H": 18, "I": 12, "J": 14, "K": 18},
        "Neighborhood Analysis": {"A": 35, "B": 14, "C": 14, "D": 14, "E": 12, "F": 16, "G": 18, "H": 18, "I": 12, "J": 14, "K": 18},
        "Economic Indicators": {"A": 38, "B": 22, "C": 60},
        "Crime & Safety": {"A": 50, "B": 18, "C": 22, "D": 60},
        "Investment Recommendations": {"A": 16, "B": 16, "C": 18, "D": 14, "E": 60},
        "Data Sources": {"A": 38, "B": 38, "C": 22, "D": 60},
    }
    for sheet_name, widths in column_widths.items():
        if sheet_name in wb.sheetnames:
            s = wb[sheet_name]
            for col, w in widths.items():
                s.column_dimensions[col].width = w

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")

    # Summary print
    five = [r["zip_code"] for r in zip_rows if score_zip(
        r.get("total_inv_trans_6mo") or 0, r.get("median_days_on_market"), r.get("median_home_value")
    ) == "★★★★★"]
    four = [r["zip_code"] for r in zip_rows if score_zip(
        r.get("total_inv_trans_6mo") or 0, r.get("median_days_on_market"), r.get("median_home_value")
    ) == "★★★★☆"]
    print(f"\n5-star ZIPs ({len(five)}): {five}")
    print(f"4-star ZIPs ({len(four)}): {four}")
    print(f"Total ZIPs: {len(zip_rows)} | Total Neighborhoods: {len(nbhd_rows)}")
    print(f"Total 6-mo investor transactions: {sum((r.get('total_inv_trans_6mo') or 0) for r in zip_rows)}")


if __name__ == "__main__":
    main()
