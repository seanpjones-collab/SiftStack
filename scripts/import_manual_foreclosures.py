"""One-shot importer for Sean's manual pre-automation foreclosure pulls.

Reads:
  - output/reports_smoke/Stark_County_Foreclosures_2026.xlsx          (full year)
  - output/reports_smoke/Stark_County_Foreclosures_2026-04-07_to_2026-04-14.xlsx
  - output/reports_smoke/Summit_County_Foreclosures_April2026.csv
(the "Stark_County_Foreclosures_2026 (2).xlsx" file is a superseded
 version of the full-year file with no address column — skipped.)

Dedupes by case number within each county. Excludes addresses already in
the 4/17+ automated backfill (e.g. 410 Patterson Ave was in both).
Runs Smarty address standardization. Emits Sift-ready DataSift CSVs.

Usage:
    python scripts/import_manual_foreclosures.py

Outputs to output/manual_import/:
    stark_foreclosures_manual_datasift.csv
    summit_foreclosures_manual_datasift.csv
"""
from __future__ import annotations

import csv
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

# Make src/ importable
REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import config  # noqa: E402,F401  (loads .env)
from notice_parser import NoticeData  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger("import_manual")


# ── Address normalization (for de-dup against scraped data) ─────────

_ABBREV = {
    "ROAD": "RD", "STREET": "ST", "AVENUE": "AVE", "DRIVE": "DR",
    "LANE": "LN", "COURT": "CT", "CIRCLE": "CIR", "PLACE": "PL",
    "BOULEVARD": "BLVD", "TERRACE": "TER",
}


def norm_addr(s: str) -> str:
    if not s:
        return ""
    s = str(s).upper().strip()
    s = s.split(",")[0].strip()  # street portion only
    s = re.sub(r"\s+(APT|UNIT|STE|SUITE|#)\s*\S*$", "", s)
    for full, abbr in _ABBREV.items():
        s = re.sub(r"\b" + full + r"\b", abbr, s)
    return re.sub(r"\s+", " ", s).strip()


# ── Name splitting ──────────────────────────────────────────────────


def split_defendant(raw: str) -> tuple[str, str]:
    """Parse defendant name into (first, last).

    Handles:
      "SMITH, JOHN Q"         → ("John Q", "Smith")
      "HOLCOMB AUDREY D"      → ("Audrey D", "Holcomb")  (Summit CSV format)
      "WILBURN"               → ("", "Wilburn")          (last name only)
      "JOHN SMITH"            → ("John", "Smith")        (rare)
    """
    raw = (raw or "").strip()
    if not raw:
        return "", ""
    # Drop parens + entity suffixes
    raw = re.sub(r"\([^)]+\)", "", raw).strip()
    raw = re.sub(r",?\s*(ET\s*AL|JR|SR|II|III|IV|TRUSTEE|TRS?)\.?$", "",
                 raw, flags=re.IGNORECASE).strip()
    raw = re.sub(r"[,.]$", "", raw).strip()

    # "LAST, FIRST MIDDLE" — comma marker
    if "," in raw:
        last, _, rest = raw.partition(",")
        return rest.strip().title(), last.strip().title()

    parts = raw.split()
    if len(parts) == 1:
        return "", parts[0].title()
    # Two or more — heuristic: if ALL-CAPS then "LAST FIRST MIDDLE" (Stark + Summit clerk format)
    # Otherwise "FIRST LAST"
    if raw == raw.upper():
        last = parts[0]
        first = " ".join(parts[1:])
        return first.title(), last.title()
    return parts[0].title(), " ".join(parts[1:]).title()


# ── Parsers ─────────────────────────────────────────────────────────


def parse_full_address(addr: str) -> tuple[str, str, str, str]:
    """Split "123 MAIN ST, CANTON, OH 44709" → (street, city, state, zip).
    Returns ("", "", "", "") if unparseable.
    """
    if not addr:
        return ("", "", "", "")
    addr = str(addr).strip()
    # Try "street, city, ST zip" format
    m = re.match(
        r"^(.+?),\s*(.+?),\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?\s*$",
        addr, re.IGNORECASE,
    )
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).upper(), m.group(4)
    # Fallback: "street, city, ST, zip"
    m = re.match(
        r"^(.+?),\s*(.+?),\s*([A-Z]{2}),?\s*(\d{5})(?:-\d{4})?\s*$",
        addr, re.IGNORECASE,
    )
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).upper(), m.group(4)
    return (addr, "", "", "")


def load_stark_xlsx(path: Path, case_seen: set) -> list[NoticeData]:
    """Load a Stark foreclosure xlsx → NoticeData list. Updates case_seen in place."""
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[NoticeData] = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        hdr = [str(c or "").strip() for c in rows[0]]
        col = {h: i for i, h in enumerate(hdr)}

        # Accept both "Case Number" and "Case #". Can't use `a or b` since
        # a valid column index of 0 is falsy and would fall through.
        def first(*keys):
            for k in keys:
                if k in col:
                    return col[k]
            return None
        case_idx      = first("Case Number", "Case #")
        addr_idx      = first("Property Address")
        name_idx      = first("Defendant / Homeowner", "Defendant/Homeowner")
        date_idx      = first("File Date")
        type_idx      = first("Type")
        plaintiff_idx = first("Plaintiff / Lender", "Plaintiff/Lender")
        court_idx     = first("Court")
        link_idx      = first("CJIS Search Link", "CJIS Link")

        if case_idx is None or addr_idx is None:
            continue

        for row in rows[1:]:
            if not row or len(row) <= addr_idx:
                continue
            case_no = str(row[case_idx] or "").strip()
            raw_addr = str(row[addr_idx] or "").strip() if addr_idx < len(row) else ""
            if not case_no or not raw_addr:
                continue
            if case_no in case_seen:
                continue
            case_seen.add(case_no)

            street, city, state, zip_code = parse_full_address(raw_addr)
            if not street or not re.search(r"\d", street):
                continue

            def_name = str(row[name_idx] or "").strip() if name_idx is not None and name_idx < len(row) else ""
            first, last = split_defendant(def_name)

            date_str = ""
            if date_idx is not None and date_idx < len(row) and row[date_idx]:
                dv = row[date_idx]
                date_str = dv.strftime("%Y-%m-%d") if hasattr(dv, "strftime") else str(dv).strip()[:10]

            ftype = str(row[type_idx] or "").strip() if type_idx is not None and type_idx < len(row) else ""
            plaintiff = str(row[plaintiff_idx] or "").strip() if plaintiff_idx is not None and plaintiff_idx < len(row) else ""
            court = str(row[court_idx] or "").strip() if court_idx is not None and court_idx < len(row) else ""
            link = str(row[link_idx] or "").strip() if link_idx is not None and link_idx < len(row) else ""

            n = NoticeData()
            n.address = street
            n.city = city.title() if city else ""
            n.state = "OH"
            n.zip = zip_code
            # _build_row calls _split_name(owner_name) — no owner_first field
            n.owner_name = f"{first} {last}".strip()
            # For foreclosures, mailing = property (owner-occupied assumption)
            n.owner_street = street
            n.owner_city = city.title() if city else ""
            n.owner_state = "OH"
            n.owner_zip = zip_code
            n.notice_type = "foreclosure"
            n.county = "Stark"
            n.date_added = date_str
            n.case_number = case_no
            n.source_url = link
            n.raw_text = (
                f"[manual_import] case={case_no} court={court} type={ftype} "
                f"date={date_str} plaintiff={plaintiff}"
            )
            out.append(n)
        break  # only first sheet
    return out


def load_summit_csv(path: Path, exclude_norm_addrs: set[str]) -> list[NoticeData]:
    """Load Summit foreclosure CSV. Skip rows whose normalized street matches
    `exclude_norm_addrs` (already captured by the automated scrape)."""
    out: list[NoticeData] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            street = (row.get("Street Address") or "").strip()
            if not street:
                continue
            if norm_addr(street) in exclude_norm_addrs:
                logger.info("skip overlap with scrape: %s", street)
                continue

            def_name = (row.get("Defendant Name") or "").strip()
            first, last = split_defendant(def_name)

            case_no = (row.get("Case Number") or "").strip()
            date_str = (row.get("Filing Date") or "").strip()
            # "04/01/2026" → "2026-04-01"
            m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", date_str)
            if m:
                date_str = f"{m.group(3)}-{m.group(1)}-{m.group(2)}"

            ftype = (row.get("Foreclosure Type") or "").strip()
            plaintiff = (row.get("Plaintiff / Lender") or "").strip()
            url = (row.get("Case URL") or "").strip()
            city = (row.get("City") or "").strip().title()
            state = (row.get("State") or "OH").strip().upper()
            zip_code = (row.get("ZIP") or "").strip()

            n = NoticeData()
            n.address = street
            n.city = city
            n.state = state
            n.zip = zip_code
            # _build_row calls _split_name(owner_name) — no owner_first field
            n.owner_name = f"{first} {last}".strip()
            n.owner_street = street
            n.owner_city = city
            n.owner_state = state
            n.owner_zip = zip_code
            n.notice_type = "foreclosure"
            n.county = "Summit"
            n.date_added = date_str
            n.case_number = case_no
            n.source_url = url
            n.raw_text = (
                f"[manual_import] case={case_no} type={ftype} "
                f"date={date_str} plaintiff={plaintiff}"
            )
            out.append(n)
    return out


# ── Main ────────────────────────────────────────────────────────────


def main() -> int:
    smoke_dir = REPO / "output" / "reports_smoke"
    out_dir = REPO / "output" / "manual_import"
    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Load Stark (2 useful xlsx files) ──
    stark_seen: set[str] = set()
    stark: list[NoticeData] = []
    stark += load_stark_xlsx(
        smoke_dir / "Stark_County_Foreclosures_2026.xlsx", stark_seen,
    )
    stark += load_stark_xlsx(
        smoke_dir / "Stark_County_Foreclosures_2026-04-07_to_2026-04-14.xlsx",
        stark_seen,
    )
    logger.info("Stark: %d NoticeData objects (deduped by case#)", len(stark))

    # Dedupe stark addresses against our automated scrape
    scraped_stark_addrs: set[str] = set()
    scraped_csv = REPO / "output" / "1.1.16_stark_backfill" / "datasift_dms.csv"
    if scraped_csv.exists():
        with open(scraped_csv, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                a = norm_addr(row.get("Property Street Address") or "")
                if a:
                    scraped_stark_addrs.add(a)
    before = len(stark)
    stark = [n for n in stark if norm_addr(n.address) not in scraped_stark_addrs]
    logger.info("Stark after overlap dedup vs backfill: %d (dropped %d)",
                len(stark), before - len(stark))

    # ── Load Summit (1 csv) ──
    scraped_summit_addrs: set[str] = set()
    scraped_csv = REPO / "output" / "1.1.16_summit_backfill" / "datasift_dms.csv"
    if scraped_csv.exists():
        with open(scraped_csv, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                a = norm_addr(row.get("Property Street Address") or "")
                if a:
                    scraped_summit_addrs.add(a)
    summit = load_summit_csv(
        smoke_dir / "Summit_County_Foreclosures_April2026.csv",
        scraped_summit_addrs,
    )
    logger.info("Summit: %d NoticeData objects (after overlap dedup)", len(summit))

    # ── Smarty address standardization ──
    if config.SMARTY_AUTH_ID and config.SMARTY_AUTH_TOKEN:
        try:
            from address_standardizer import standardize_addresses
            logger.info("Running Smarty on %d records...", len(stark) + len(summit))
            standardize_addresses(stark,  config.SMARTY_AUTH_ID, config.SMARTY_AUTH_TOKEN)
            standardize_addresses(summit, config.SMARTY_AUTH_ID, config.SMARTY_AUTH_TOKEN)
            # Smarty updates owner_street/city/state/zip too when mailing matches property
            for n in stark + summit:
                if n.dpv_match_code == "Y" and not n.owner_street:
                    n.owner_street = n.address
                    n.owner_city = n.city
                    n.owner_state = n.state
                    n.owner_zip = n.zip
            s_confirmed = sum(1 for n in stark if n.dpv_match_code == "Y")
            u_confirmed = sum(1 for n in summit if n.dpv_match_code == "Y")
            logger.info("Smarty USPS-confirmed: Stark %d/%d, Summit %d/%d",
                        s_confirmed, len(stark), u_confirmed, len(summit))
        except ImportError:
            logger.warning("smartystreets-python-sdk not installed — skipping")
        except Exception as e:
            logger.warning("Smarty failed: %s — continuing without standardization", e)
    else:
        logger.warning("SMARTY creds not set in .env — skipping standardization")

    # ── Emit DataSift CSVs ──
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    stark_out = out_dir / f"stark_foreclosures_manual_{ts}.csv"
    summit_out = out_dir / f"summit_foreclosures_manual_{ts}.csv"

    # write_datasift_csv writes to config.OUTPUT_DIR / filename so we
    # override the destination path manually after writing.
    # Simpler: just invoke the row-builder directly.
    from datasift_formatter import _build_row, DATASIFT_COLUMNS, _build_dm_notes

    def write_as_sift(records: list[NoticeData], path: Path) -> None:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=DATASIFT_COLUMNS)
            w.writeheader()
            for n in records:
                w.writerow(_build_row(n, notes_override=_build_dm_notes(n)))

    write_as_sift(stark, stark_out)
    write_as_sift(summit, summit_out)
    logger.info("Stark CSV:  %s  (%d rows)", stark_out, len(stark))
    logger.info("Summit CSV: %s  (%d rows)", summit_out, len(summit))
    logger.info("Upload both at app.reisift.io → Upload File → Add Data")
    return 0


if __name__ == "__main__":
    sys.exit(main())
