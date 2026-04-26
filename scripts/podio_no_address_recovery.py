"""Recover property addresses for the 111 no-address Podio records.

Cross-references the no-address CSV against any of these sources (whichever
exist on disk at runtime — works with one, both, or neither):

  tmp/podio_migration/sift_records_export.csv      — manual Sift export
  tmp/podio_migration/smrtdialer_lists/*.csv|*.xlsx — your smrtDialer source lists

Theory: when a list contact called back without being pushed to Podio first,
Podio created a record from the smrtPhone comm-log sync — phone + name only,
no address. The original smrtDialer source list is the lookup table.

Match key: phone number (normalized to 10-digit US local format on both sides).
For each Podio no-address record, walks every populated phone (Phone 1..9) and
returns the first source that has a property address. Files are searched in
order: Sift export first (highest data quality), then smrtDialer files.

Outputs:
  output/podio_migration/podio_no_address_recovered.csv  — addresses found, ready to upload
  output/podio_migration/podio_no_address_pending.csv    — still need calls per Kylie
  output/podio_migration/podio_no_address_recovery_log.csv — per-record match details

Usage:
    python scripts/podio_no_address_recovery.py
"""
from __future__ import annotations

import csv
import logging
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

from datasift_formatter import DATASIFT_COLUMNS  # noqa: E402

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger("recovery")

NO_ADDR_CSV = REPO / "output" / "podio_migration" / "podio_no_address.csv"
SIFT_EXPORT = REPO / "tmp" / "podio_migration" / "sift_records_export.csv"
SMRT_DIALER_DIR = REPO / "tmp" / "podio_migration" / "smrtdialer_lists"

OUTPUT_DIR = REPO / "output" / "podio_migration"
RECOVERED_CSV = OUTPUT_DIR / "podio_no_address_recovered.csv"
PENDING_CSV = OUTPUT_DIR / "podio_no_address_pending.csv"
LOG_CSV = OUTPUT_DIR / "podio_no_address_recovery_log.csv"


# ── Phone normalization ─────────────────────────────────────────────


def norm_phone(raw) -> str:
    """Normalize phone to 10-digit US format. Returns '' if unrecoverable."""
    if not raw:
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits if len(digits) == 10 else ""


# ── File loaders (CSV or XLSX) ──────────────────────────────────────


def read_table(path: Path) -> tuple[list[str], list[dict]]:
    """Read a CSV or XLSX file into (fieldnames, list-of-dicts)."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return (reader.fieldnames or [], list(reader))
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]  # first sheet
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return ([], [])
        headers = [str(h) if h is not None else "" for h in rows[0]]
        out = []
        for r in rows[1:]:
            out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
        return (headers, out)
    raise ValueError(f"Unsupported file type: {path}")


# ── Phone-index builder (works on any phone+address table) ──────────


def _detect_columns(fieldnames: list[str]) -> dict:
    """Find address + phone columns by case-insensitive substring match."""
    def find_col(*needles: str) -> str | None:
        for needle in needles:
            for col in fieldnames:
                if needle.lower() in col.lower():
                    return col
        return None

    # "property" needles first so we don't accidentally grab "Mailing address"
    # when both "Property address" and "Mailing address" columns exist (Sift export).
    street_col = find_col("property street", "property address",
                          "street address", "address line 1", "address1",
                          "street", "address")
    city_col   = find_col("property city", "city")
    state_col  = find_col("property state", "state")
    zip_col    = find_col("property zip", "zip code", "zip", "postal")
    # Phone columns — column name must FULLY match one of these patterns.
    # Excludes Phone1_DNC / Phone1_Score / Phone1_Type / Phone1_Last_Seen
    # metadata that DataAtomic-style exports put alongside Phone1_Number.
    phone_patterns = [
        r"^phone\s*\d*$",                       # Phone, Phone 1, Phone1
        r"^phone[\s_]?\d*[\s_]?number$",        # Phone_Number, Phone1_Number, Phone 1 Number
        r"^primary\s*phone$",
        r"^mobile(?:\s*phone)?(?:\s*\d+)?$",    # Mobile, Mobile Phone, Mobile 1
        r"^cell(?:\s*phone)?(?:\s*\d+)?$",
        r".*\bph\s*number$",                    # "PR Ph Number", "Attorney Ph Number"
        r".*\bphone\s*number$",                 # "PR Phone Number"
        r".*\bphone\s*\d*$",                    # "Owner Phone 1"
    ]
    phone_re = re.compile("|".join(f"(?:{p})" for p in phone_patterns),
                          re.IGNORECASE)
    phone_cols = [c for c in fieldnames if c and phone_re.match(c)]
    if not phone_cols:
        # Last-resort fallback for unusual exports
        phone_cols = [c for c in fieldnames if "phone" in c.lower()]
    return {
        "street": street_col, "city": city_col, "state": state_col,
        "zip": zip_col, "phones": phone_cols,
    }


def build_index_from_rows(
    rows: list[dict], fieldnames: list[str], source_label: str,
) -> dict[str, dict]:
    """Build phone → address index from a list of dict rows."""
    cols = _detect_columns(fieldnames)
    logger.info(
        "[%s] %d rows. cols: street=%r city=%r state=%r zip=%r phones=%d",
        source_label, len(rows), cols["street"], cols["city"],
        cols["state"], cols["zip"], len(cols["phones"]),
    )
    if not cols["phones"]:
        logger.warning("[%s] no phone columns detected — index empty", source_label)
        return {}

    def cell(row, col):
        if not col:
            return ""
        v = row.get(col, "")
        if v is None:
            return ""
        return str(v).strip()

    index: dict[str, dict] = {}
    for row in rows:
        addr = {
            "street": cell(row, cols["street"]),
            "city":   cell(row, cols["city"]),
            "state":  cell(row, cols["state"]),
            "zip":    cell(row, cols["zip"]),
            "_source_label": source_label,
        }
        if not (addr["street"] or addr["city"] or addr["zip"]):
            continue
        for pcol in cols["phones"]:
            np = norm_phone(row.get(pcol))
            if np:
                index.setdefault(np, addr)
    logger.info("[%s] indexed %d unique phone→address mappings", source_label, len(index))
    return index


def load_sift_index() -> dict[str, dict]:
    if not SIFT_EXPORT.exists():
        logger.info("[Sift] not found, skipping: %s", SIFT_EXPORT)
        return {}
    fieldnames, rows = read_table(SIFT_EXPORT)
    return build_index_from_rows(rows, fieldnames, "Sift")


def load_smrtdialer_index() -> dict[str, dict]:
    """Walk the smrtdialer_lists/ folder, build one merged phone→address index.

    First file's match wins on collision (filename-sorted alphabetically).
    """
    if not SMRT_DIALER_DIR.exists():
        logger.info("[smrtDialer] folder not found, skipping: %s", SMRT_DIALER_DIR)
        return {}
    files = sorted(
        f for f in SMRT_DIALER_DIR.iterdir()
        if f.is_file() and f.suffix.lower() in (".csv", ".xlsx", ".xlsm")
    )
    if not files:
        logger.info("[smrtDialer] no CSV/XLSX files in %s", SMRT_DIALER_DIR)
        return {}

    merged: dict[str, dict] = {}
    for path in files:
        try:
            fieldnames, rows = read_table(path)
        except Exception as e:
            logger.warning("[smrtDialer] failed to read %s: %s", path.name, e)
            continue
        idx = build_index_from_rows(rows, fieldnames, f"smrtDialer/{path.name}")
        for phone, addr in idx.items():
            merged.setdefault(phone, addr)
    logger.info("[smrtDialer] merged %d unique phone→address mappings across %d files",
                len(merged), len(files))
    return merged


# ── Main ────────────────────────────────────────────────────────────


def main():
    if not NO_ADDR_CSV.exists():
        logger.error("Input not found: %s — run podio_to_datasift.py first", NO_ADDR_CSV)
        sys.exit(1)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    sift_index = load_sift_index()
    smrt_index = load_smrtdialer_index()

    if not sift_index and not smrt_index:
        logger.error(
            "No source data available. Drop one or both of these into place:\n"
            "  %s  (Sift Records export → save as CSV)\n"
            "  %s  (any smrtDialer list CSVs/XLSXs — drop multiple files in this folder)\n"
            "Then re-run.",
            SIFT_EXPORT, SMRT_DIALER_DIR,
        )
        sys.exit(1)

    # Iterate the no-address records
    with open(NO_ADDR_CSV, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        no_addr_rows = list(reader)
        no_addr_fields = reader.fieldnames or DATASIFT_COLUMNS

    recovered: list[dict] = []
    pending: list[dict] = []
    log_rows: list[dict] = []

    for row in no_addr_rows:
        # Walk all phone fields on the no-address record
        candidate_phones = []
        for i in range(1, 10):
            np = norm_phone(row.get(f"Phone {i}"))
            if np and np not in candidate_phones:
                candidate_phones.append(np)

        match = None
        match_source = ""
        match_phone = ""
        for np in candidate_phones:
            if np in sift_index:
                match = sift_index[np]
                match_source = "Sift"
                match_phone = np
                break
            if np in smrt_index:
                match = smrt_index[np]
                match_source = match["_source_label"]  # e.g., "smrtDialer/list1.csv"
                match_phone = np
                break

        log_rows.append({
            "Owner First Name": row.get("Owner First Name", ""),
            "Owner Last Name": row.get("Owner Last Name", ""),
            "Phones Tried": ",".join(candidate_phones),
            "Match Source": match_source,
            "Matched Phone": match_phone,
            "Recovered Street": match["street"] if match else "",
            "Recovered City":   match["city"] if match else "",
            "Recovered State":  match["state"] if match else "",
            "Recovered ZIP":    match["zip"] if match else "",
        })

        if match and match["street"]:
            row = dict(row)
            row["Property Street Address"] = match["street"]
            row["Property City"] = match["city"]
            row["Property State"] = match["state"]
            row["Property ZIP Code"] = match["zip"]
            # Tag origin: address-recovered-sift  OR  address-recovered-smrtdialer
            origin = "sift" if match_source == "Sift" else "smrtdialer"
            existing_tags = row.get("Tags", "")
            recovery_tag = f"address-recovered-{origin}"
            row["Tags"] = (existing_tags + "," + recovery_tag) if existing_tags else recovery_tag
            recovered.append(row)
        else:
            pending.append(row)

    def write_csv(path: Path, rows: list[dict], fieldnames):
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                writer.writerow({k: r.get(k, "") for k in fieldnames})

    write_csv(RECOVERED_CSV, recovered, no_addr_fields)
    write_csv(PENDING_CSV, pending, no_addr_fields)
    write_csv(LOG_CSV, log_rows, list(log_rows[0].keys()) if log_rows else [])

    logger.info("─" * 60)
    logger.info("Total no-address records:   %d", len(no_addr_rows))
    logger.info("Recovered with address:     %d  → %s", len(recovered), RECOVERED_CSV.name)
    logger.info("Still pending (no match):   %d  → %s", len(pending), PENDING_CSV.name)
    logger.info("Per-record log:             %s", LOG_CSV.name)
    logger.info("Sources used: Sift=%d phones, smrtDialer=%d phones",
                len(sift_index), len(smrt_index))

    # Recovery breakdown by source
    if log_rows:
        from collections import Counter
        source_counts = Counter(r["Match Source"] for r in log_rows if r["Match Source"])
        if source_counts:
            logger.info("Match breakdown:")
            for src, count in source_counts.most_common():
                logger.info("  %s: %d", src, count)


if __name__ == "__main__":
    main()
