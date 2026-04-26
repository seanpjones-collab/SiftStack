"""Podio Seller Leads → DataSift batched CSVs.

Reads tmp/podio_migration/Seller Leads - Last view used.xlsx and emits:
  output/podio_migration/podio_batch1_hot.csv             (HOT temperature)
  output/podio_migration/podio_batch2_active.csv          (active pipeline + followup + sold)
  output/podio_migration/podio_batch3_dispositioned.csv   (Dead/DNC/Not Owner/Lost)
  output/podio_migration/podio_no_address.csv             (no Property Address Map — recovery target)

Per Kylie's guidance:
  - Every record tagged "podio leads"
  - Property Status left blank — tags drive SiftLine phasing later
  - First/last name split, address parsed into street/city/state/zip
  - Mailing fields blank (Sift enrichment fills later)
  - LLCs/Trusts handled per existing SiftStack pattern: name fields blank,
    full entity name in Notes, tagged entity_owned

Tags on every record:
  - "podio leads"
  - "podio-status:<Lead Status>"
  - "podio-temp:<Temperature>" if Temperature set
  - "podio-source:<Lead Source>" if Lead Source set
  - "entity_owned" if name parses as business entity

Notes (single field) concatenates these Podio fields with labels:
  Follow Up Notes, Reason for Selling, Time Frame to Sell, Asking Price,
  Anyone Living in House, Repairs/Maintenance Needed, Property Condition,
  Years Owned, Property Paid Off, original entity name.

Usage:
    python scripts/podio_to_datasift.py
"""
from __future__ import annotations

import csv
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

from datasift_formatter import (  # noqa: E402
    DATASIFT_COLUMNS,
    _clean_and_split_name,
    _is_entity_name,
)
from openpyxl import load_workbook  # noqa: E402

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger("podio_to_datasift")

INPUT_XLSX = REPO / "tmp" / "podio_migration" / "Seller Leads - Last view used.xlsx"
OUTPUT_DIR = REPO / "output" / "podio_migration"

# Lead Status values that route to the dispositioned (Batch 3) bucket.
DISPOSITIONED_STATUSES = {
    "Dead",
    "DNC List",
    "Not Owner/ Bad Name/ Bad Phone Number",
    "Lost Deal",
}

# Notes field assembly: (Podio column header, label in Notes)
NOTES_SOURCE_FIELDS = [
    ("Follow Up Notes", "Follow-up Notes"),
    ("Reason_for_Selling", "Reason for Selling"),
    ("Time_Frame_To_Sell", "Time Frame to Sell"),
    ("Asking_Price", "Asking Price"),
    ("Anyone_Living_In_House", "Occupancy"),
    ("Repairs_Maintenance_Needed", "Repairs Needed"),
    ("How_Long_Owned_Property", "Years Owned"),
    ("Property Paid off?", "Paid Off"),
    ("Property Condition", "Condition"),
    ("Misc Property Notes", "Misc"),
    ("Details (Condition, Reason for Selling, How soon they want to sell)", "Details"),
]


# Street suffixes used to find the street/city boundary in space-separated
# addresses like "2464 Sherwin Dr Twinsburg OH 44087".
_STREET_SUFFIX_WORDS = {
    "st", "street", "ave", "avenue", "rd", "road", "dr", "drive",
    "ln", "lane", "ct", "court", "cir", "circle", "pl", "place",
    "blvd", "boulevard", "ter", "terrace", "way", "pkwy", "parkway",
    "hwy", "highway", "trl", "trail", "pt", "point", "row",
}
_DIRECTIONAL_WORDS = {"n", "s", "e", "w", "ne", "nw", "se", "sw"}

_FULL_STATE_NAMES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY",
}


def _split_street_city(body: str) -> tuple[str, str]:
    """Split "<street...> <city...>" on a street suffix (+optional directional)."""
    words = body.split()
    if len(words) < 2:
        return (body, "")
    suffix_idx = -1
    for i, w in enumerate(words):
        if w.lower().rstrip(".,") in _STREET_SUFFIX_WORDS:
            suffix_idx = i
    if suffix_idx == -1:
        return (body, "")
    end_idx = suffix_idx
    if (end_idx + 1 < len(words)
            and words[end_idx + 1].lower().rstrip(".,") in _DIRECTIONAL_WORDS):
        end_idx += 1
    if end_idx + 1 >= len(words):
        return (body, "")
    street = " ".join(words[: end_idx + 1])
    city = " ".join(words[end_idx + 1 :])
    return (street, city)


def parse_full_address(addr) -> tuple[str, str, str, str]:
    """Parse Podio Property Address Map → (street, city, state, zip).

    Handles three formats observed in Sean's Podio export:
      - "301 Melbourne Ave, Akron, OH 44313, USA"        (Google-style)
      - "73010 Old Twenty One Rd, Kimbolton, Ohio, 43749" (full state name)
      - "2464 Sherwin Dr Twinsburg OH 44087"             (no commas)
    """
    if not addr:
        return ("", "", "", "")
    s = str(addr).strip()
    # Drop trailing ", USA"
    s = re.sub(r",?\s*USA\s*$", "", s, flags=re.IGNORECASE)
    # Normalize full state names ("Ohio" → "OH", "Tennessee" → "TN", ...)
    for full, abbr in _FULL_STATE_NAMES.items():
        s = re.sub(rf",\s*{full}\b", f", {abbr}", s, flags=re.IGNORECASE)

    # Comma-separated with zip
    m = re.match(
        r"^(.+?),\s*(.+?),\s*([A-Z]{2}),?\s+(\d{5})(?:-\d{4})?\s*$",
        s, re.IGNORECASE,
    )
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).upper(), m.group(4)

    # Comma-separated, no zip ("East Blvd, Cleveland, OH")
    m = re.match(r"^(.+?),\s*(.+?),\s*([A-Z]{2})\s*$", s, re.IGNORECASE)
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).upper(), ""

    # Space-separated: "<street...> <city...> ST 12345"
    m = re.match(
        r"^(.+?)\s+([A-Z]{2})\s+(\d{5})(?:-\d{4})?\s*$", s, re.IGNORECASE,
    )
    if m:
        body = m.group(1).strip()
        state = m.group(2).upper()
        zip_code = m.group(3)
        street, city = _split_street_city(body)
        if city:
            return (street, city, state, zip_code)

    logger.debug("Unparseable address: %r", addr)
    return (s, "", "", "")


def fmt_date(val) -> str:
    """Convert datetime to M/D/YYYY for DataSift."""
    if not val:
        return ""
    if isinstance(val, datetime):
        return f"{val.month}/{val.day}/{val.year}"
    return str(val).strip()


def fmt_phone(val) -> str:
    """Normalize a phone field. Podio stores digits-only; pass through if set."""
    if not val:
        return ""
    s = str(val).strip()
    return s if s else ""


def collect_phones(row: dict) -> list[str]:
    """Build ordered phone list from Podio fields. Skips fax variants."""
    order = [
        "Seller Phone - Mobile",
        "Seller Phone - Work",
        "Seller Phone - Home",
        "Seller Phone - Main",
        "Seller Phone - Other",
    ]
    seen = set()
    out = []
    for key in order:
        p = fmt_phone(row.get(key))
        if p and p not in seen:
            seen.add(p)
            out.append(p)
    return out


def collect_emails(row: dict) -> list[str]:
    order = [
        "Seller Email - Work",
        "Seller Email - Home",
        "Seller Email - Other",
    ]
    seen = set()
    out = []
    for key in order:
        v = row.get(key)
        if v:
            v = str(v).strip()
            if v and v not in seen:
                seen.add(v)
                out.append(v)
    return out


def build_notes(row: dict, entity_name: str = "") -> str:
    """Concatenate Podio narrative fields into a single Notes string."""
    parts = []
    if entity_name:
        parts.append(f"Entity: {entity_name}")
    for col, label in NOTES_SOURCE_FIELDS:
        v = row.get(col)
        if v is None or v == "":
            continue
        s = str(v).strip()
        if s:
            parts.append(f"{label}: {s}")
    return " | ".join(parts)


def build_tags(row: dict, is_entity: bool) -> str:
    tags = ["podio leads"]
    status = row.get("Lead Status")
    if status:
        tags.append(f"podio-status:{str(status).strip()}")
    temp = row.get("Temperature")
    if temp:
        tags.append(f"podio-temp:{str(temp).strip()}")
    source = row.get(" Lead Source")  # note leading space in Podio header
    if source:
        tags.append(f"podio-source:{str(source).strip()}")
    if is_entity:
        tags.append("entity_owned")
    return ",".join(tags)


def assign_batch(row: dict, has_address: bool) -> str:
    """Return batch label: 'hot', 'active', 'dispositioned', or 'no_address'."""
    if not has_address:
        return "no_address"
    temp = (row.get("Temperature") or "").strip()
    status = (row.get("Lead Status") or "").strip()
    if temp == "HOT":
        return "hot"
    if status in DISPOSITIONED_STATUSES:
        return "dispositioned"
    return "active"


def build_csv_row(row: dict, batch: str) -> dict:
    """Build a DataSift CSV row dict from a Podio row dict."""
    raw_addr = row.get("Property Address Map") or ""
    street, city, state, zip_code = parse_full_address(raw_addr)

    raw_name = (row.get("Seller Name") or "").strip()
    is_entity = _is_entity_name(raw_name) if raw_name else False
    if is_entity:
        first, last = "", ""
        entity_name_for_notes = raw_name
    else:
        first, last = _clean_and_split_name(raw_name)
        entity_name_for_notes = ""

    phones = collect_phones(row)
    emails = collect_emails(row)
    notes = build_notes(row, entity_name=entity_name_for_notes)
    tags = build_tags(row, is_entity)
    date_added = fmt_date(row.get("Date Lead Came In"))

    list_name = {
        "hot":            "Podio Migration - Hot",
        "active":         "Podio Migration - Active",
        "dispositioned":  "Podio Migration - Dispositioned",
        "no_address":     "Podio Migration - No Address",
    }[batch]

    out = {col: "" for col in DATASIFT_COLUMNS}
    out["Property Street Address"] = street
    out["Property City"] = city
    out["Property State"] = state
    out["Property ZIP Code"] = zip_code
    out["Owner First Name"] = first
    out["Owner Last Name"] = last
    # Mailing fields intentionally blank — Kylie's instruction
    for i, p in enumerate(phones[:9], start=1):
        out[f"Phone {i}"] = p
    for i, e in enumerate(emails[:5], start=1):
        out[f"Email {i}"] = e
    out["Tags"] = tags
    out["Lists"] = list_name
    out["Notes"] = notes
    out["Date Added"] = date_added
    return out


def main():
    if not INPUT_XLSX.exists():
        logger.error("Input file not found: %s", INPUT_XLSX)
        sys.exit(1)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    ws = wb["Seller Leads"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = rows[1:]
    wb.close()

    # Bucket rows by batch
    buckets: dict[str, list[dict]] = {
        "hot": [], "active": [], "dispositioned": [], "no_address": [],
    }
    entity_count = 0
    unparseable_addr = 0

    for raw_row in data:
        row = {h: v for h, v in zip(headers, raw_row)}
        addr = row.get("Property Address Map")
        has_address = bool(addr and str(addr).strip())
        batch = assign_batch(row, has_address)
        csv_row = build_csv_row(row, batch)
        buckets[batch].append(csv_row)

        if csv_row["Tags"] and "entity_owned" in csv_row["Tags"]:
            entity_count += 1
        if has_address and not csv_row["Property City"]:
            unparseable_addr += 1

    # Write batch CSVs
    file_map = {
        "hot": "podio_batch1_hot.csv",
        "active": "podio_batch2_active.csv",
        "dispositioned": "podio_batch3_dispositioned.csv",
        "no_address": "podio_no_address.csv",
    }
    for batch, fname in file_map.items():
        path = OUTPUT_DIR / fname
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=DATASIFT_COLUMNS)
            writer.writeheader()
            for r in buckets[batch]:
                writer.writerow(r)
        logger.info("Wrote %d records → %s", len(buckets[batch]), path)

    # Summary
    total = sum(len(v) for v in buckets.values())
    logger.info("─" * 60)
    logger.info("Total Podio rows processed: %d", total)
    logger.info("  Batch 1 (HOT):           %d", len(buckets["hot"]))
    logger.info("  Batch 2 (active):        %d", len(buckets["active"]))
    logger.info("  Batch 3 (dispositioned): %d", len(buckets["dispositioned"]))
    logger.info("  No-address (recovery):   %d", len(buckets["no_address"]))
    logger.info("Entities (LLCs etc.):     %d", entity_count)
    if unparseable_addr:
        logger.warning("Addresses that didn't parse cleanly: %d", unparseable_addr)


if __name__ == "__main__":
    main()
