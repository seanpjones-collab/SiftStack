"""Recalibrate Summit + Stark market research workbooks to FRED=55 (March 2026
reading) so star ratings are apples-to-apples across all three OH counties.

Updates only the DOM vs NATIONAL and WHOLESALING SCORE columns on the
ZIP Code Analysis and Neighborhood Analysis sheets. Narrative sheets
(Executive Summary, Economic, Crime, Recommendations, Data Sources) are
left untouched — they'll stay correct directionally; the FRED baseline
note in Data Sources will be updated.

Usage:
    python scripts/recalibrate_market_research.py
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
FILES = [
    REPO / "output" / "ftm_county_research" / "Summit_County_OH_Market_Research.xlsx",
    REPO / "output" / "ftm_county_research" / "Stark_County_OH_Market_Research.xlsx",
]
NEW_FRED = 55  # March 2026 FRED MEDDAYONMARUS reading


def score(trans, dom, val) -> str:
    """Same algorithm used in Cuyahoga script — keeps tier definitions
    consistent across all three workbooks."""
    if not trans or dom is None or dom == "—":
        return "★☆☆☆☆"
    try:
        dom_int = int(dom)
        trans_int = int(trans)
    except (ValueError, TypeError):
        return "★☆☆☆☆"
    val_num = None
    if val and val != "—":
        try:
            val_num = float(re.sub(r"[$,]", "", str(val)))
        except (ValueError, TypeError):
            pass
    delta = dom_int - NEW_FRED
    if delta >= 10:
        return "★☆☆☆☆"
    if delta > 0:
        return "★★☆☆☆"
    if trans_int >= 30 and delta <= -10 and val_num and val_num < 350_000:
        return "★★★★★"
    if trans_int >= 20 and val_num and val_num < 400_000:
        return "★★★★☆"
    if trans_int >= 10:
        return "★★★☆☆"
    return "★★☆☆☆"


def recalibrate_sheet(ws) -> dict[str, int]:
    """Update DOM vs NATIONAL (col F) and WHOLESALING SCORE (col K) in place.
    Skips rows 1 (title) and 2 (header). Returns counts of each star tier."""
    tier_counts = {"★★★★★": 0, "★★★★☆": 0, "★★★☆☆": 0, "★★☆☆☆": 0, "★☆☆☆☆": 0}
    for row_idx in range(3, ws.max_row + 1):
        trans = ws.cell(row=row_idx, column=2).value
        dom = ws.cell(row=row_idx, column=5).value
        val = ws.cell(row=row_idx, column=7).value
        if dom is None or dom == "—":
            new_delta = "—"
        else:
            try:
                new_delta = f"{int(dom) - NEW_FRED:+d}"
            except (ValueError, TypeError):
                new_delta = "—"
        ws.cell(row=row_idx, column=6, value=new_delta)
        new_score = score(trans, dom, val)
        ws.cell(row=row_idx, column=11, value=new_score)
        tier_counts[new_score] += 1
    return tier_counts


def update_data_sources(ws) -> None:
    """Update the FRED baseline reference on the Data Sources sheet."""
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if not v or not isinstance(v, str):
                continue
            # DOM vs National formula row
            if "FRED National Baseline" in v and ("days" in v or "Day" in v):
                ws.cell(
                    row=row_idx, column=col_idx,
                    value=f"Median DOM − FRED National Baseline ({NEW_FRED} days, March 2026)",
                )
            elif v.endswith("days)") and "national" in v.lower() and "FRED" in v:
                ws.cell(
                    row=row_idx, column=col_idx,
                    value=f"FRED National DOM Baseline ({NEW_FRED} days, March 2026)",
                )


def main() -> None:
    for path in FILES:
        if not path.exists():
            print(f"MISSING: {path}")
            continue
        print(f"\n=== {path.name} ===")
        wb = load_workbook(path)
        for sheet_name in ("ZIP Code Analysis", "Neighborhood Analysis"):
            if sheet_name not in wb.sheetnames:
                print(f"  [skip] no {sheet_name} sheet")
                continue
            counts = recalibrate_sheet(wb[sheet_name])
            print(f"  {sheet_name}: " + " | ".join(
                f"{star} {n}" for star, n in counts.items()
            ))
        if "Data Sources" in wb.sheetnames:
            update_data_sources(wb["Data Sources"])
            print(f"  Data Sources: FRED baseline note updated to {NEW_FRED} days")
        wb.save(path)
        print(f"  Saved.")

    print("\nRecalibration complete. Re-extract hot zips with the same query as before.")


if __name__ == "__main__":
    main()
